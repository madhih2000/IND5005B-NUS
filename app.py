from io import BytesIO
import streamlit as st
import streamlit_ext as ste
from st_aggrid import AgGrid, GridOptionsBuilder
import pandas as pd
import numpy as np
import os
import tempfile
import scipy.stats as stats
from scipy.stats import norm, poisson, nbinom, gamma, weibull_min, lognorm, expon, beta, kstest, anderson
import zipfile

from openpyxl import load_workbook
from openpyxl import load_workbook
from PIL import Image
from openpyxl.drawing.image import Image as XLImage  
from openpyxl.utils.dataframe import dataframe_to_rows

import plotly.graph_objects as go
import plotly.subplots as sp
import plotly.express as px
import plotly.figure_factory as ff
import plotly.io as pio

from utils import *
import consumption_utils
import order_placement_utils
import goods_receipt_utils
import forecast_models
import lead_time_analysis
import DES
import llm_reasoning
import waterfall_analysis

# Set the page config with the title centered
st.set_page_config(page_title="Micron SupplySense", layout="wide")

# Center title at the top
st.markdown(
    """
    <h1 style="text-align: center; color: #4B9CD3;">Micron SupplySense</h1>
    """, 
    unsafe_allow_html=True
)

# Create a sidebar for navigation (for a dashboard-style layout)
tabs = st.sidebar.radio("Select an Analysis Type:", ["Material Consumption Analysis", "Order Placement Analysis", "Goods Receipt Analysis","Lead Time Analysis", "Forecast Demand", "Inventory Simulation", "Waterfall Analysis"])

if tabs == "Material Consumption Analysis":
    st.title("Material Consumption Analysis")
    # Upload Excel files
    uploaded_file = st.file_uploader("Upload Consumption Excel File for Analysis", type=["xlsx"])

    if uploaded_file:
        df = load_data_consumption(uploaded_file)  # Read the file

        # Check if 'Material Group' column exists
        if 'Material Group' in df.columns:
            unique_groups = df['Material Group'].astype(str).unique()  # Get unique values

            for group in unique_groups:
                st.subheader(f"Material Group {group} Analysis")

                # Filter data for the current Material Group
                df_filtered = df[df['Material Group'].astype(str) == str(group)]

                # Run analysis functions in the correct order
                df_more_filtered,top_n = consumption_utils.overall_consumption_patterns(df_filtered)
                consumption_utils.outlier_detection(df_more_filtered, top_n)
                consumption_utils.shelf_life_analysis(df_filtered)
                #consumption_utils.vendor_consumption_analysis(df_filtered)
                #consumption_utils.location_consumption_analysis(df_filtered)
                #consumption_utils.batch_variability_analysis(df_filtered)
                #consumption_utils.combined_analysis(df_filtered)
                consumption_utils.specific_material_analysis(df_filtered)

        else:
            st.error("The uploaded file does not contain a 'Material Group' column. Please check the file format.")

elif tabs == "Order Placement Analysis":
    st.title("Order Placement Analysis")

    # File uploader
    uploaded_file = st.file_uploader("Upload Order Placement Excel File for Analysis", type="xlsx")

    if uploaded_file:
        df = order_placement_utils.preprocess_order_data(uploaded_file)  # Read the file

        # Check if 'Material Group' column exists
        if 'Material Group' in df.columns:
            unique_groups = df['Material Group'].astype(str).unique()  # Get unique values

            for group in unique_groups:
                st.subheader(f"Material Group {group} Analysis")

                # Filter data for the current Material Group
                df_filtered = df[df['Material Group'].astype(str) == str(group)]

                # Call the analysis functions
                df_more_filtered,top_n = order_placement_utils.overall_orderplacement_patterns(df_filtered)
                order_placement_utils.outlier_detection(df_more_filtered, top_n)
                # order_placement_utils.overall_order_patterns(df_filtered)
                # order_placement_utils.outlier_detection(df_filtered)
                # order_placement_utils.vendor_order_analysis(df_filtered)
                # order_placement_utils.order_trends_over_time(df_filtered)
                # order_placement_utils.monthly_order_patterns(df_filtered)
                # order_placement_utils.vendor_material_analysis(df_filtered)
                # order_placement_utils.plant_order_analysis(df_filtered)
                # order_placement_utils.purchasing_document_analysis(df_filtered)
                # order_placement_utils.order_quantity_distribution(df_filtered)
                # order_placement_utils.material_vendor_analysis(df_filtered)
                # order_placement_utils.supplier_order_analysis(df_filtered)
                # order_placement_utils.material_plant_analysis(df_filtered)
                # order_placement_utils.abc_analysis(df_filtered)
                order_placement_utils.specific_material_analysis(df_filtered)


    else:
        st.write("Please upload an Excel file to begin the analysis.")


elif tabs == "Goods Receipt Analysis":
    st.title("Goods Receipt Analysis")

    # File uploader
    uploaded_file = st.file_uploader("Upload Goods Receipt Excel File for Analysis", type="xlsx")

    if uploaded_file:
        df = load_data_GR(uploaded_file)  # Read the file

        # Check if 'Material Group' column exists
        if 'Material Group' in df.columns:
            unique_groups = df['Material Group'].astype(str).unique()  # Get unique values

            for group in unique_groups:
                st.subheader(f"Material Group {group} Analysis")

                # Filter data for the current Material Group
                df_filtered = df[df['Material Group'].astype(str) == str(group)]

                # Call the analysis functions
                df_more_filtered,top_n = goods_receipt_utils.overall_GR_patterns(df_filtered)
                goods_receipt_utils.outlier_detection(df_more_filtered, top_n)
                goods_receipt_utils.specific_material_analysis(df_filtered)


    else:
        st.write("Please upload an Excel file to begin the analysis.")

elif tabs == "Forecast Demand":
    st.title("Forecast Model")

    # File uploader
    uploaded_file = st.file_uploader("Upload Consumption Data Excel File for Analysis", type="xlsx")

    if uploaded_file:
        df = load_forecast_consumption_data(uploaded_file)

        if df is not None and 'Material Number' in df.columns:
            material_numbers = df['Material Number'].unique()
            selected_material_number = st.selectbox("Select Material Number", material_numbers)
            filtered_df = df[df['Material Number'] == selected_material_number].copy()

            model_choice = st.selectbox("Select Model", ["XGBoost", "ARIMA"])
            forecast_weeks = st.number_input("Forecast Weeks", min_value=1, value=6)
            seasonality = st.selectbox("Seasonality", ["Yes", "No"])

            # Run forecast only if button is clicked
            if st.button("Run Forecast"):
                with st.spinner("Running forecast, please wait..."):
                    if model_choice == "XGBoost":
                        forecast_results, plt = forecast_models.forecast_weekly_consumption_xgboost_plotly(
                            filtered_df, forecast_weeks_ahead=forecast_weeks, seasonality=seasonality)
                    else:
                        forecast_results, plt = forecast_models.forecast_weekly_consumption_arima_plotly(
                            filtered_df, forecast_weeks_ahead=forecast_weeks, seasonality=seasonality)

                    # Store in session_state
                    st.session_state.forecast_results = forecast_results
                    st.session_state.plot = plt
                    st.session_state.params_df = pd.DataFrame({
                        "Selected Material Number": [selected_material_number],
                        "Model Used": [model_choice],
                        "Forecast Weeks": [forecast_weeks],
                        "Seasonality": [seasonality],
                    })
                    st.session_state.filename = f"forecast_{str(selected_material_number).replace(' ', '_')}_{model_choice}_{forecast_weeks}w_{seasonality.lower()}.xlsx"

        # Display forecast and download only if session state has results
        if "forecast_results" in st.session_state:
            st.write(f"{st.session_state.params_df['Model Used'][0]} Forecast Results:")
            st.plotly_chart(st.session_state.plot)
            st.write(st.session_state.forecast_results)

            # Create in-memory Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write DataFrames
                st.session_state.forecast_results.to_excel(writer, sheet_name='Forecast Results', index=False)
                st.session_state.params_df.to_excel(writer, sheet_name='Parameters', index=False)
                
                # Save the workbook object for further editing
                #writer.book  # Trigger openpyxl backend
                #writer.sheets  # Trigger sheet creation

            # Save and reopen to insert image
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.create_sheet("Forecast Plot")

            # Save plot to image in memory
            img_bytes = pio.to_image(st.session_state.plot, format='png', width=1000, height=600)
            img_stream = BytesIO(img_bytes)

            # Load image into openpyxl Image
            pil_img = Image.open(img_stream)
            img_for_excel = XLImage(img_stream)
            img_for_excel.anchor = 'A1'  # Position on sheet

            # Add image to sheet
            ws.add_image(img_for_excel)

            # Save final workbook to new BytesIO
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            # Download button
            ste.download_button(
                label="Download Forecast Excel",
                data=final_output,  
                file_name=st.session_state.filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        elif df is not None:
            if 'Material Number' not in df.columns:
                st.error("The uploaded file does not contain a 'Material Number' column.")

elif tabs == "Lead Time Analysis":
    st.title("")

    # File uploader
    uploaded_file_op = st.file_uploader("Upload Order Placement Excel File for Analysis", type="xlsx")
    uploaded_file_gr = st.file_uploader("Upload Goods Received Excel File for Analysis", type="xlsx")
    uploaded_file_sr = st.file_uploader("Upload Latest Shortage Report", type="xlsx")

    if uploaded_file_op and uploaded_file_gr and uploaded_file_sr:
        with st.spinner("Processing lead time analysis..."):
            op_df = pd.read_excel(uploaded_file_op)
            gr_df = pd.read_excel(uploaded_file_gr)
            shortage_df = pd.read_excel(uploaded_file_sr)

            matched, unmatched_op, unmatched_gr = lead_time_analysis.process_dataframes(op_df, gr_df)
            calculated_df = lead_time_analysis.calculate_actual_lead_time(matched)
            final_df = lead_time_analysis.calculate_lead_time_summary_v2(shortage_df)
            final_result = lead_time_analysis.calculate_lead_time_differences(final_df, calculated_df)

            # Option to pick Supplier from matched
            if 'Supplier' in final_result.columns:
                suppliers = final_result['Supplier'].unique().tolist()
                selected_supplier = st.selectbox("Select Supplier (Optional)", ["All"] + suppliers)

                if selected_supplier != "All":
                    filtered_final_result = final_result[final_result['Supplier'] == selected_supplier]
                else:
                    filtered_final_result = final_result
            else:
                filtered_final_result = final_result
                st.write("Supplier column not found in matched data.")

            # Option to filter based on 'Pstng Date' (Received Date) from gr_df
            if 'Pstng Date' in gr_df.columns:
                min_date = gr_df['Pstng Date'].min().date()
                max_date = gr_df['Pstng Date'].max().date()
                selected_date_range = st.date_input("Filter By Goods Received Date", (min_date, max_date))

                if len(selected_date_range) == 2:
                    start_date, end_date = selected_date_range
                    gr_df_filtered = gr_df[
                        (gr_df['Pstng Date'].dt.date >= start_date) &
                        (gr_df['Pstng Date'].dt.date <= end_date)
                    ]
                    #recalculate matched, and other dataframes based on filtered gr_df
                    matched, unmatched_op, unmatched_gr = lead_time_analysis.process_dataframes(op_df, gr_df_filtered)
                    calculated_df = lead_time_analysis.calculate_actual_lead_time(matched)
                    final_result = lead_time_analysis.calculate_lead_time_differences(final_df, calculated_df)
                    if 'Supplier' in final_result.columns:
                        if selected_supplier != "All":
                            filtered_final_result = final_result[final_result['Supplier'] == selected_supplier]
                        else:
                            filtered_final_result = final_result
                            fig5, fig6, fig7 = lead_time_analysis.plot_supplier_lead_time_analysis(filtered_final_result)
                    else:
                        filtered_final_result = final_result
                        st.write("Supplier column not found in matched data.")

            else:
                st.write("Pstng Date column not found in Goods Received data.")

            # Call the updated Plotly version of your function
            fig1, fig2, fig3, fig4 = lead_time_analysis.analyze_and_plot_lead_time_differences_plotly(filtered_final_result)


        st.success("Lead Time Analysis Completed ✅")
        st.write("### Material Level Lead Time Analysis Results:")
        st.plotly_chart(fig1, use_container_width=True)
        st.plotly_chart(fig2, use_container_width=True)
        st.plotly_chart(fig3, use_container_width=True)
        st.plotly_chart(fig4, use_container_width=True)

        if selected_supplier == "All":
            st.markdown(
                """
                <style>
                    .gradient-line {
                        height: 10px;
                        background: linear-gradient(to right, #0000FF, #008000);
                        border: none;
                        margin: 10px 0;
                    }
                </style>
                <hr class="gradient-line">
                """,
                unsafe_allow_html=True
            )
            st.write("### Supplier Level Lead Time Analysis Results:")
            st.plotly_chart(fig5, use_container_width=True)
            st.plotly_chart(fig6, use_container_width=True)
            st.plotly_chart(fig7, use_container_width=True)




    else:
        st.write("Please upload all Excel files to begin the analysis.")

elif tabs == "Waterfall Analysis":
    st.title("Waterfall Analysis")
    uploaded_file = st.file_uploader("Upload Zip file with Shortage Data (e.g. WW1.xlsx to WW52.xlsx)", type=["zip"])
    uploaded_file_op = st.file_uploader("Upload Order Placement Excel File", type="xlsx")
    uploaded_file_gr = st.file_uploader("Upload Goods Receipt Excel File", type="xlsx")
    uploaded_file_cons = st.file_uploader("Upload Consumption Excel File", type=["xlsx"])

    if uploaded_file and uploaded_file_op and uploaded_file_gr and uploaded_file_cons:
        op_df = pd.read_excel(uploaded_file_op)
        gr_df = pd.read_excel(uploaded_file_gr)
        cons_df = load_data_consumption_waterfall(uploaded_file_cons)


        with tempfile.TemporaryDirectory() as tmp_dir:
            zip_path = os.path.join(tmp_dir, "data.zip")
            zip_filename = uploaded_file.name
            zip_ext = os.path.splitext(zip_filename)[0]
            with open(zip_path, "wb") as f:
                f.write(uploaded_file.getvalue())

            try:
                with zipfile.ZipFile(zip_path, "r") as zip_ref:
                    zip_ref.extractall(tmp_dir)
            except zipfile.BadZipFile:
                st.error("❌ The uploaded file is not a valid ZIP.")

            folder_path_zip = os.path.join(tmp_dir, zip_ext)

            if not os.path.exists(folder_path_zip):
                st.error(f"❌ Error: Folder '{folder_path_zip}' not found.")

            xlsx_files = []
            for root, dirs, files in os.walk(folder_path_zip):
                for file in files:
                    if file.endswith(".xlsx"):
                        xlsx_files.append(os.path.relpath(os.path.join(root, file), folder_path_zip))

            if not xlsx_files:
                st.error("❌ Error: No XLSX files found in the ZIP.")

            # Get just the filenames
            all_files = [os.path.basename(f) for f in xlsx_files]

            # Optional: Store full paths if needed later
            file_paths = {os.path.basename(f): os.path.join(folder_path_zip, f) for f in xlsx_files}

            # Expected 52 week files
            expected_files = [f"WW{i}.xlsx" for i in range(1, 53)]

            # Check which expected files are missing
            missing_files = [f for f in expected_files if f not in all_files]
            if missing_files:
                st.error(f"❌ Missing files: {', '.join(missing_files)}")

            # User inputs
            start_week_str = st.selectbox("Select Start Week", [f"WW{i}" for i in range(2,53)])  # WW2 to WW52
            if start_week_str:
                start_week = int(start_week_str.replace("WW", ""))
                # Get the weekly file path starting from start_week
                selected_weeks = f"{start_week_str}.xlsx"

                df = pd.read_excel(file_paths[selected_weeks])

                if not df.empty:
                    # Material dropdown
                    material_options = sorted(df["Material Number"].dropna().unique())
                    material_number = st.selectbox("Select Material Number", material_options)

                    if material_number:
                        plant_options = sorted(df[df["Material Number"] == material_number]["Plant"].dropna().unique())
                        plant = st.selectbox("Select Plant", plant_options)

                        if plant:
                            site_options = sorted(
                                df[
                                    (df["Material Number"] == material_number) &
                                    (df["Plant"] == plant)
                                ]["Site"].dropna().unique()
                            )
                            site = st.selectbox("Select Site", site_options)

                            num_weeks = st.number_input("Number of Weeks", min_value=1, max_value=26, value=12)

                            # Submit button to extract & display data
                            if st.button("Run Waterfall Analysis"):
                                with st.spinner("Running Analysis..."):
                                    result_df, lead_value = waterfall_analysis.extract_and_aggregate_weekly_data(
                                        folder_path_zip, material_number, plant, site, start_week, cons_df, int(num_weeks)
                                    )

                                    if result_df is not None and not result_df.empty:
                                        st.success("✅ Data extracted successfully!")
                                        st.markdown(
                                            """
                                            <style>
                                                .gradient-line {
                                                    height: 10px;
                                                    background: linear-gradient(to right, #0000FF, #008000);
                                                    border: none;
                                                    margin: 10px 0;
                                                }
                                            </style>
                                            <hr class="gradient-line">
                                            """,
                                            unsafe_allow_html=True
                                        )
                                        st.subheader("Waterfall Chart")

                                        st.dataframe(result_df)

                                        st.subheader("PO Summary Table")

                                        PO_df = merged_order_gr_PO_analysis(op_df, gr_df)
                                        PO_df_filtered = PO_df[
                                            (PO_df['Material Number'] == material_number) &
                                            (PO_df['Plant'] == plant) &
                                            (PO_df['Site'] == site)
                                        ].reset_index(drop=True).sort_values(by=['Order Date', 'Purchasing Document']).reset_index(drop=True)

                                        if PO_df_filtered.empty:
                                            st.warning("No matching PO records found for the selected Material Number, Plant, and Site.")

                                        else:
                                            st.dataframe(PO_df_filtered)

                                        st.header("Root Cause Analysis")

                                        # RCA Condition 1
                                        try:
                                            st.subheader('Validation of WoS levels against PO coverage')
                                            scen_1_df_output = waterfall_analysis.scenario_1(result_df, PO_df_filtered)
                                            styled_df = waterfall_analysis.style_dataframe(scen_1_df_output)
                                            st.dataframe(styled_df, use_container_width=True)
                                            analysis_1 = llm_reasoning.explain_scenario_1_with_groq(scen_1_df_output)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 1: {e}")

                                        try:
                                            st.subheader('Comparison of Actual & Predicted WoS')
                                            wos_list, analysis_plot, comparison_table = waterfall_analysis.plot_stock_prediction_plotly(result_df, start_week, lead_value, num_weeks)
                                            st.plotly_chart(analysis_plot)
                                            st.write("Forecast Accuracy Validation Table")
                                            st.dataframe(comparison_table)
                                            analysis_2 = llm_reasoning.explain_scenario_2_with_groq(comparison_table)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 2: {e}")

                                        # RCA Condition 3
                                        try:
                                            st.subheader('Scenario 3 - Visualization of Inventory Levels and PO Adjustment Strategies')                                        
                                            st.write("PO Timing Analysis")
                                            po_analysis_output = waterfall_analysis.scenario_3(result_df, PO_df_filtered, lead_value)
                                            st.dataframe(po_analysis_output)
                                            analysis_3 = llm_reasoning.explain_scenario_3_with_groq(comparison_table)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 3: {e}")

                                        # RCA Condition 4
                                        try:
                                            st.subheader('Changes in Lead Time')
                                            condition4 = waterfall_analysis.lead_time_check(result_df)
                                            st.dataframe(condition4)
                                            analysis_4 = llm_reasoning.explain_scenario_4_with_groq(condition4)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 4: {e}")

                                        # RCA Condition 5
                                        try:
                                            st.subheader('Inspection of Demand w/o Buffer Patterns')
                                            condition5 = waterfall_analysis.analyze_week_to_week_demand_changes(result_df, lead_time=lead_value)
                                            st.dataframe(condition5)
                                            weekly_demand_summary = waterfall_analysis.calculate_weekly_demand_summary(condition5)
                                            st.info("""
                                                ### **Weekly Demand Variability Summary — Column Guide**

                                                This table helps explain how stable or unstable the weekly demand has been, based on past updates.

                                                - **Week**: The calendar week (e.g., WW12) we're analyzing.

                                                - **SD (Standard Deviation)**: Tells us how much the demand numbers changed during the week.  
                                                ▸ A **higher number** means demand was bouncing around a lot — it wasn’t consistent.

                                                - **CV (Coefficient of Variation)**: This compares the change (SD) to the average demand.  
                                                ▸ Helpful when comparing products with very different demand levels — it adjusts for size.

                                                - **Spike**: Number of times during the week that demand jumped by **more than 10 units**.  
                                                ▸ Think of this as mini "demand surges".

                                                - **Drop**: Number of times during the week that demand **fell by more than 10 units**.  
                                                ▸ These are significant sudden drops.

                                                - **Sudden % Spike**: How many times demand jumped by **30% or more** in a single update.  
                                                ▸ Shows extreme upward swings, even if units are small.

                                                - **Sudden % Drop**: How many times demand fell by **30% or more** in a single update.  
                                                ▸ Highlights steep declines that may need attention.

                                                - **Avg Abs WoW Change**: The **average amount** demand changed from one update to the next (ignores direction).  
                                                ▸ This gives a sense of how much demand moved around week-to-week.

                                                - **Irregularity Score**: A **summary score** that rolls up everything above — big jumps, swings, and inconsistency.  
                                                ▸ Higher scores = more erratic or unstable demand behavior.
                                                """)
                                            st.write("Weekly Demand Variability Summary")
                                            st.dataframe(weekly_demand_summary)
                                            analysis_5 = llm_reasoning.explain_scenario_5_with_groq(condition5, weekly_demand_summary)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 5: {e}")

                                        # RCA Condition 6
                                        try:
                                            st.subheader('Inspection of Consumption Patterns')
                                            consumption_vals, fig, comparison_df = waterfall_analysis.plot_consumption_vs_demand_plotly(result_df)
                                            st.plotly_chart(fig)
                                            st.write("Analysis of Consumption Against Planned Demand")
                                            st.dataframe(comparison_df)

                                            condition6 = waterfall_analysis.scenario_6(result_df, PO_df_filtered)
                                            cons_reported_act = condition6[["Snapshot Week", "Consumption (Waterfall)", "Consumption (Calc)"]]
                                            cons_reported_act["Abs Diff"] = (cons_reported_act["Consumption (Waterfall)"] - cons_reported_act["Consumption (Calc)"]).abs()

                                            def flag_discrepancy(row, threshold=0.15):
                                                calc = row["Consumption (Calc)"]
                                                rep = row["Consumption (Waterfall)"]
                                                if calc == 0 and rep == 0:
                                                    return ""
                                                if calc == 0 or rep == 0:
                                                    return "⚠️ Zero mismatch"
                                                ratio = rep / calc
                                                if ratio > 1 + threshold:
                                                    return f"⚠️ Over by {round((ratio - 1) * 100)}%"
                                                elif ratio < 1 - threshold:
                                                    return f"⚠️ Under by {round((1 - ratio) * 100)}%"
                                                return ""

                                            cons_reported_act["Flag"] = cons_reported_act.apply(flag_discrepancy, axis=1)

                                            st.write("Consumption Comparison (Reported vs Calculated)")
                                            st.dataframe(cons_reported_act)

                                            with st.expander("Show End-to-End Inventory and Consumption Tracking"):
                                                st.write("End-to-End Inventory and Consumption Tracking")
                                                st.dataframe(condition6)

                                            analysis_6 = llm_reasoning.explain_scenario_6_with_groq(condition6)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 6: {e}")

                                        try:
                                            st.subheader('Supply vs Goods Receipt Gap Analysis')
                                            condition6["Snapshot Week Num"] = condition6["Snapshot Week"].str.replace("WW", "").astype(int)
                                            condition_7 = condition6[["Snapshot Week", "Snapshot Week Num", "Supply (Waterfall)", "PO GR Quantity"]].rename(columns={"PO GR Quantity": "GR Quantity"})
                                            # Merge on week number
                                            merged_df = pd.merge(condition_7, PO_df_filtered[["GR WW", "Purchasing Document"]], 
                                                                left_on='Snapshot Week Num', right_on='GR WW', how='left')
                                            # Drop the temporary merge key if you don't want it in the final result
                                            merged_df.drop(columns=["Snapshot Week Num"], inplace=True)
                                            # Group and aggregate POs by week
                                            summary_df = merged_df.groupby("Snapshot Week").agg({
                                                "Supply (Waterfall)": "first",
                                                "GR Quantity": "first",
                                                "Purchasing Document": lambda x: list(x.dropna())
                                            }).reset_index()

                                            # Apply the discrepancy function here
                                            summary_df[['Discrepancy_Flag', 'Discrepancy_Detail', 'Abs_Difference']] = summary_df.apply(waterfall_analysis.analyze_discrepancy_scen_7, axis=1)
                                            cols = ['Snapshot Week', 'Supply (Waterfall)', 'GR Quantity', 'Abs_Difference',
                                                    'Purchasing Document', 'Discrepancy_Flag', 'Discrepancy_Detail']

                                            summary_df = summary_df[cols]
                                            st.dataframe(summary_df)

                                            analysis_7 = llm_reasoning.explain_scenario_7_with_groq(summary_df)
                                        except Exception as e:
                                            st.error(f"Error in Scenario 7: {e}")

                                        rca_final = llm_reasoning.explain_waterfall_chart_with_groq(result_df, analysis_1, analysis_2, analysis_3, analysis_4, analysis_5, analysis_6, analysis_7)

                                        # Download button
                                        output = BytesIO()
                                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                            # 1. Write Waterfall Chart sheet (unstyled for now)
                                            result_df.to_excel(writer, sheet_name="Waterfall Chart", index=False)
                                            # 2. Parameters
                                            pd.DataFrame({
                                                "Parameter": ["Start Week", "Material Number", "Plant", "Site", "Number of Weeks"],
                                                "Value": [start_week_str, material_number, plant, site, num_weeks]
                                            }).to_excel(writer, sheet_name="Chosen Parameters", index=False)

                                            # PO Summary table
                                            PO_df_filtered.to_excel(writer, sheet_name="PO Summary", index=False)

                                            # RCA Summary
                                            writer.book.create_sheet("RCA Summary")
                                            rca_final_sheet = writer.sheets["RCA Summary"]
                                            rca_final_sheet.append(["Root Cause Analysis (Summary)"])
                                            write_analysis_block(rca_final_sheet, rca_final)

                                            # RCA Condition Sheets
                                            writer.book.create_sheet("RCA Scenario 1")
                                            cond1_sheet = writer.sheets["RCA Scenario 1"]
                                            cond1_sheet.append(["Scenario 1 - PO Coverage is Inadequate"])
                                            for r in dataframe_to_rows(scen_1_df_output, index=False, header=True):
                                                cond1_sheet.append(r)

                                            write_analysis_block(cond1_sheet, analysis_1)

                                            # --- Scenario 2: Comparison of Actual & Predicted Weeks of Supply ---
                                            cond2_sheet = writer.book.create_sheet("RCA Scenario 2")
                                            cond2_sheet.append(["Scenario 2 - Comparison of Actual & Predicted Weeks of Supply"])

                                            # Insert Plotly Forecast Accuracy Chart
                                            try:
                                                img_bytes = analysis_plot.to_image(format="png", width=1000, height=600)
                                                img_stream = BytesIO(img_bytes)
                                                img = XLImage(img_stream)
                                                img.anchor = "A3"  # Place image starting at cell A3
                                                cond2_sheet.add_image(img)
                                            except Exception as e:
                                                cond2_sheet.append([f"[Error inserting forecast chart image: {e}]"])

                                            # Start writing tabular data below image
                                            current_row = 35

                                            # Forecast Accuracy Table
                                            cond2_sheet.cell(row=current_row, column=1, value="Forecast Accuracy Validation Table")
                                            current_row += 1
                                            if 'comparison_table' in locals():
                                                for r in dataframe_to_rows(comparison_table, index=False, header=True):
                                                    for col_idx, cell_value in enumerate(r, start=1):
                                                        cond2_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                                                    current_row += 1

                                            # Analysis text (LLM reasoning)
                                            current_row += 2
                                            write_analysis_block(cond2_sheet, locals().get('analysis_2', ""), start_row=current_row)

                                            # --- Scenario 3: Inventory & PO Adjustment Analysis ---
                                            cond3_sheet = writer.book.create_sheet("RCA Scenario 3")
                                            cond3_sheet.append(["Scenario 3 - Inventory Analysis and Optimized PO Adjustment Strategies"])

                                            current_row = 3
                                            if 'po_analysis_output' in locals():
                                                for r in dataframe_to_rows(po_analysis_output, index=False, header=True):
                                                    for col_idx, cell_value in enumerate(r, start=1):
                                                        cond3_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                                                    current_row += 1

                                            # Analysis block (if available)
                                            current_row += 2
                                            write_analysis_block(cond3_sheet, locals().get('analysis_3', ""), start_row=current_row)

                                            cond4_sheet = writer.book.create_sheet("RCA Scenario 4")
                                            cond4_sheet.append(sanitize_row(["Scenario 4 - Longer Delivery Lead Time"]))
                                            for r in dataframe_to_rows(condition4, index=False, header=True):
                                                cond4_sheet.append(sanitize_row(r))
                                            write_analysis_block(cond4_sheet, analysis_4)

                                            cond5_sheet = writer.book.create_sheet("RCA Scenario 5")
                                            cond5_sheet.append(["Scenario 5 - Irregular Demand w/o Buffer Patterns"])
                                            for r in dataframe_to_rows(condition5, index=False, header=True):
                                                cond5_sheet.append(r)
                                            write_analysis_block(cond5_sheet, analysis_5)

                                            cond6_sheet = writer.book.create_sheet("RCA Scenario 6")
                                            cond6_sheet.append(["Scenario 6 - Irregular Consumption Patterns"])
                                            try:
                                                # Convert plotly figure to PNG bytes in memory
                                                img_bytes = fig.to_image(format="png", width=1000, height=600)
                                                img_stream = BytesIO(img_bytes)

                                                # Create openpyxl image from bytes stream
                                                img = XLImage(img_stream)
                                                img.anchor = "A3"  # Position image starting at cell A3
                                                cond6_sheet.add_image(img)

                                            except Exception as e:
                                                cond6_sheet.append([f"[Error inserting image: {e}]"])
                                            # Start writing from row 30
                                            start_row = 35
                                            current_row = start_row

                                            # Write comparison_df
                                            cond6_sheet.cell(row=current_row, column=1, value="Analysis of Consumption Against Planned Demand")
                                            current_row += 1
                                            for r in dataframe_to_rows(comparison_df, index=False, header=True):
                                                for col_idx, cell_value in enumerate(r, start=1):
                                                    cond6_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                                                current_row += 1

                                            # Leave a blank row
                                            current_row += 1

                                            # Write Reported vs Calculated Consumption Comparison table
                                            cond6_sheet.cell(row=current_row, column=1, value="Reported vs Calculated Consumption Comparison")
                                            current_row += 1

                                            if 'cons_reported_act' in locals():
                                                for row in dataframe_to_rows(cons_reported_act, index=False, header=True):
                                                    for col_idx, cell_value in enumerate(row, start=1):
                                                        cond6_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                                                    current_row += 1

                                            # Leave a blank row
                                            current_row += 1

                                            # Write condition6 table
                                            cond6_sheet.cell(row=current_row, column=1, value="End-to-End Inventory and Consumption Tracking")
                                            current_row += 1
                                            for r in dataframe_to_rows(condition6, index=False, header=True):
                                                for col_idx, cell_value in enumerate(r, start=1):
                                                    cond6_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                                                current_row += 1

                                            # Leave another blank row
                                            current_row += 1

                                            # Write the analysis block below
                                            write_analysis_block(cond6_sheet, analysis_6, start_row=current_row)

                                            # --- Scenario 7: Supply vs Goods Receipt Gap Analysis ---
                                            cond7_sheet = writer.book.create_sheet("RCA Scenario 7")
                                            cond7_sheet.append(["Scenario 7 - Supply vs Goods Receipt Gap Analysis"])

                                            # Start writing summary_df table starting from row 3
                                            current_row = 3
                                            if 'summary_df' in locals() and "Purchasing Document" in summary_df.columns:
                                                summary_df["Purchasing Document"] = summary_df["Purchasing Document"].apply(
                                                    lambda x: ", ".join(map(str, x)) if isinstance(x, list) else str(x)
                                                )
                                                for row in dataframe_to_rows(summary_df, index=False, header=True):
                                                    for col_idx, cell_value in enumerate(row, start=1):
                                                        cond7_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                                                    current_row += 1

                                            # Leave a gap and write the analysis block (LLM reasoning)
                                            current_row += 2
                                            write_analysis_block(cond7_sheet, locals().get('analysis_7', ""), start_row=current_row)

                                        output.seek(0)
                                        # Apply coloring on 'Waterfall Chart' sheet
                                        colored_output = waterfall_analysis.apply_coloring_to_output(
                                                        output,
                                                        lead_time=lead_value,
                                                        sheet_names=["Waterfall Chart", "RCA Scenario 1"]
                                                    )
                                        # Display download button
                                        ste.download_button(
                                            label="📥 Download Excel File",
                                            data=colored_output,
                                            file_name="Waterfall_RCA_Report.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                        )

                                    else:
                                        st.warning("No data returned from the extraction. The material number does not exist in prior weeks of the shortage data.")
            
elif tabs == "Inventory Simulation":
    st.title("Inventory Simulation")

    # File uploader
    uploaded_consumption = st.file_uploader("Upload Consumption File", type="xlsx")
    uploaded_goods_receipt = st.file_uploader("Upload Goods Receipt Excel File", type="xlsx")
    uploaded_order_placement = st.file_uploader("Upload Order Placement Excel File", type="xlsx")
    uploaded_merged = st.file_uploader("Upload Shortage Zip File",  type=["zip"])
               
    # Load files only when they are first uploaded
    if uploaded_consumption and "consumption_df" not in st.session_state:
        DES.load_and_store_file(uploaded_consumption, "consumption_df")

    if uploaded_goods_receipt and "gr_df" not in st.session_state:
        DES.load_and_store_file(uploaded_goods_receipt, "gr_df")

    if uploaded_order_placement and "order_df" not in st.session_state:
        DES.load_and_store_file(uploaded_order_placement, "order_df")

    if uploaded_merged and "merged_df" not in st.session_state:
        DES.load_zip_folder(uploaded_merged, "merged_df")
    
    # Access stored files without reloading
    if "consumption_df" in st.session_state:
        consumption_df = st.session_state["consumption_df"]

    if "gr_df" in st.session_state:
        gr_df = st.session_state["gr_df"]

    if "order_df" in st.session_state:
        order_df = st.session_state["order_df"]

    if "merged_df" in st.session_state:
        merged_df = st.session_state["merged_df"]

    if "consumption_df" in st.session_state and "order_df" in st.session_state and "gr_df" in st.session_state and "merged_df" in st.session_state:

        consumption_df = DES.preprocess_data_consumption(consumption_df)
 
        gr_df = DES.preprocess_data_GR(gr_df)

        order_df = DES.preprocess_data_OP(order_df)

        # Find common material numbers across all DataFrames
        common_materials = set(consumption_df['Material Number'].unique()) & \
                        set(order_df['Material Number'].unique()) & \
                        set(gr_df['Material Number'].unique()) & \
                        set(merged_df['Material Number'].unique())

        if not common_materials:
            st.warning("No common material numbers found across all DataFrames. Please check the data.")
        

        col1, col2, col3 = st.columns(3)

        # Filtering Data Before Running Simulation
        with col1:
            selected_material = st.selectbox("Select Material", list(common_materials))
            material_specific_consumption_df = consumption_df[consumption_df['Material Number'] == selected_material]
            
        with col2:
            selected_plant = st.selectbox("Select Plant", material_specific_consumption_df['Plant'].unique())

        with col3:
            selected_site = st.selectbox("Select Site", material_specific_consumption_df['Site'].unique())

        filtered_consumption = consumption_df[(consumption_df['Material Number'] == selected_material) & 
                                        (consumption_df['Plant'] == selected_plant) & 
                                        (consumption_df['Site'] == selected_site)].reset_index(drop=True)
        filtered_orders = order_df[(order_df['Material Number'] == selected_material) & 
                                (order_df['Plant'] == selected_plant)].reset_index(drop=True)
        filtered_receipts = gr_df[(gr_df['Material Number'] == selected_material) & 
                                    (gr_df['Plant'] == selected_plant) & 
                                    (gr_df['Site'] == selected_site)].reset_index(drop=True)
        filtered_merged = merged_df[(merged_df['Material Number'] == selected_material) & 
                                    (merged_df['Plant'] == selected_plant) & 
                                    (merged_df['Site'] == selected_site)].reset_index(drop=True)
        
        max_lead_time, std_lead_time, dist_name, dist_params = DES.process_lead_time(filtered_merged)

        with col1:
            num_weeks = st.number_input("Number of Simulation Weeks", min_value=1, value=52)
            st.info("Set the number of weeks for the simulation.")

            if not filtered_consumption.empty and "BUn" in filtered_consumption.columns:
                unit = filtered_consumption["BUn"].iloc[0]
            else:
                unit = ""
            initial_inventory = st.number_input(f"Initial Inventory (BUn: {unit})", min_value=10, max_value=20000, value=50)
            st.info("The starting inventory level for the simulation.")

            # Consumption Input
            consumption_type = st.radio("Consumption Type", ["Fixed", "Distribution"])
            consumption_values = []
            consumption_distribution_params = None  
            consumption_best_distribution = None
            safety_stock = 0
            if consumption_type == "Fixed":
                fixed_consumption = st.number_input("Fixed Consumption Value", min_value=0, value=10)
                consumption_values = [fixed_consumption] * num_weeks

                service_level_percentage = st.number_input("Desired Service Level (%)", min_value=1, max_value=100, value=95)
                service_level = service_level_percentage / 100.0
                std_dlt = np.sqrt(max_lead_time) * np.std(consumption_values)
                z_score = stats.norm.ppf(service_level)
                safety_stock = z_score * std_dlt
                safety_stock = int(safety_stock)
                st.success(f"Calculated Safety Stock: {safety_stock} units")

                average_consumption = np.mean(consumption_values)
                lead_time_demand = average_consumption * max_lead_time

                # 2. Calculate Reorder Point
                reorder_pt_calc = lead_time_demand + safety_stock

                reorder_point = st.number_input("Reorder Point", min_value=5, max_value=500, value=int(reorder_pt_calc))
                st.info("The inventory level at which a new order is placed.")

            else:  # Consumption Type is "Distribution"
                consumption_values = filtered_consumption.iloc[:, 4:].values.flatten()
                consumption_distribution_params, consumption_best_distribution  = DES.fit_distribution(consumption_values, "Consumption")
                mean_consumption = DES.get_mean_from_distribution(consumption_best_distribution, consumption_distribution_params)
                std_consumption = DES.get_std_from_distribution(consumption_best_distribution, consumption_distribution_params)
                if consumption_distribution_params:
                    simulated_demand = DES.simulate_demand(consumption_best_distribution, consumption_distribution_params)
                    lead_time_values = filtered_merged.filter(like="Lead Time").iloc[0].dropna().astype(float)
                    lead_time_distribution_params, lead_time_best_distribution  = DES.fit_distribution_lead_time(lead_time_values, "Lead Time")
                    simulated_lead_times = DES.simulate_demand(lead_time_best_distribution, lead_time_distribution_params)

                service_level_percentage = st.number_input("Desired Service Level (%)", min_value=1, max_value=100, value=95)
                service_level = service_level_percentage / 100.0
                # std_dlt = np.sqrt(max_lead_time) * std_consumption
                # z_score = stats.norm.ppf(service_level)
                # safety_stock = z_score * std_dlt

                simulated_stock_levels = simulated_lead_times * simulated_demand
                # Calculate safety stock as the percentile of the simulated stock levels based on the service level
                safety_stock = np.percentile(simulated_stock_levels, service_level)
                safety_stock = int(safety_stock)

                st.success(f"Calculated Safety Stock: {safety_stock} units")

                lead_time_demand = mean_consumption * max_lead_time
                # 2. Calculate Reorder Point
                reorder_pt_calc = lead_time_demand + safety_stock
                try:
                    reorder_point = st.number_input("Reorder Point", min_value=5, max_value=10000, value=int(reorder_pt_calc))
                except Exception:
                    reorder_point = st.number_input("Reorder Point", min_value=5, max_value=10000, value=100)
                st.info("The inventory level at which a new order is placed.")

        with col2:
            lead_time = st.number_input("Lead Time (weeks)", min_value=1.0, max_value=20.0, value=float(max_lead_time))
            st.info("The time (in weeks) it takes for an order to arrive after it is placed.")

            # Demand Surge Controls
            demand_surge_weeks_options = [f"WW{i+1}" for i in range(num_weeks)]
            demand_surge_weeks_input = st.multiselect("Demand Surge Weeks", demand_surge_weeks_options)
            st.info("Select the weeks where you want to simulate a sudden increase in demand.")

            min_order_qty = st.number_input("Minimum Order Quantity", min_value=1, max_value=10000, value=50)
            st.info("Select the minimum order quantity for this material number to prevent small orders during simulation.")

            # Order Quantity Input
            order_quantity_type = st.radio("Order Quantity Type", ["Fixed", "Distribution"])
            order_quantity = 0
            order_distribution_params = None
            order_distribution_best = "Fixed Distribution"
            


            if order_quantity_type == "Fixed":
                order_quantity = st.number_input("Order Quantity", min_value=10, max_value=10000, value=min_order_qty)
            else:  # Order Quantity Type is "Distribution"
                order_values = filtered_orders.iloc[:, 3:].values.flatten()
                order_distribution_params, order_distribution_best = DES.fit_distribution(order_values, "Order Quantity")

            
        with col3:
            lead_time_std_dev = st.number_input("Lead Time Std Dev (weeks)", min_value=0.0, max_value=10.0, value=float(std_lead_time))
            st.info("The standard deviation of the lead time, representing variability.")
            
            demand_surge_factor = st.number_input("Demand Surge Factor", min_value=0.5, max_value=5.0, value=2.0, step=0.1)
            st.info("Enter the factor by which demand will increase during the selected weeks. (e.g., 2.0 doubles demand)")

            N = st.number_input("Number of Monte Carlo Simulations", min_value=1, max_value=10000, value=1)
            st.info("The number of Monte Carlo simulations to run. A higher number provides more accurate results but requires more computation.")

        if st.button("Run Simulation"):
            with st.spinner("Running simulation..."):
                args = (filtered_consumption, filtered_orders, filtered_receipts, initial_inventory, reorder_point, order_quantity, lead_time, lead_time_std_dev, demand_surge_weeks_input, demand_surge_factor, consumption_distribution_params, consumption_type, consumption_best_distribution, consumption_values, num_weeks, order_distribution_params, order_distribution_best, order_quantity_type, min_order_qty)

                # Run Monte Carlo simulation
                
                all_inventory_histories, all_proactive_inventory_histories, all_stockout_weeks, all_proactive_stockout_weeks, all_wos_histories, all_proactive_wos_histories, all_consumption_histories, all_weekly_events = DES.run_monte_carlo_simulation(N, *args)
                # Compute averages
            
                avg_inventory, avg_wos, avg_consumption, stockout_frequency, avg_proactive_inventory,avg_proactive_wos, stockout_frequency_proactive = DES.compute_averages(all_inventory_histories, all_proactive_inventory_histories, all_stockout_weeks, all_proactive_stockout_weeks, all_wos_histories, all_proactive_wos_histories, all_consumption_histories)

                # # Find the representative run
                representative_index = DES.find_representative_run(all_inventory_histories, avg_inventory)

                # # Get details of the representative run
                representative_inventory, representative_inventory_proactive, representative_stockout_weeks, representative_stockout_weeks_proactive, representative_wos, representative_wos_proactive, representative_consumption, representative_weekly_events = DES.get_representative_run_details(representative_index, all_inventory_histories, all_proactive_inventory_histories, all_stockout_weeks, all_proactive_stockout_weeks, all_wos_histories, all_proactive_wos_histories, all_consumption_histories, all_weekly_events)

                # Monte Carlo Simulation (Proactive + Reactive)

                week_numbers = list(range(1, num_weeks + 1))
                # Create DataFrames for inventory, WoS, and consumption
                inventory_df = pd.DataFrame({
                    'Working Week': week_numbers,
                    'Reactive Inventory': representative_inventory,
                    'Proactive Inventory': representative_inventory_proactive
                })

                wos_df = pd.DataFrame({
                    'Working Week': week_numbers,
                    'Reactive WoS': representative_wos,
                    'Proactive WoS': representative_wos_proactive
                })

                consumption_df = pd.DataFrame({
                    'Working Week': week_numbers,
                    'Consumption': representative_consumption
                })

                # Create figure using px
                fig_inventory = px.line(
                    inventory_df, x="Working Week", y=["Reactive Inventory", "Proactive Inventory"],
                    title="Inventory Over Time", labels={"value": "Inventory", "variable": "Type"}
                )

                fig_inventory.update_xaxes(dtick=5)

                # Add shaded vertical rectangles for stockout weeks
                for week in representative_stockout_weeks:
                    fig_inventory.add_vrect(
                        x0=week - 0.5, x1=week + 0.5,  # Small padding to make it visible
                        fillcolor="red", opacity=0.25, line_width=0
                    )

                for week in representative_stockout_weeks_proactive:
                    fig_inventory.add_vrect(
                        x0=week - 0.5, x1=week + 0.5,
                        fillcolor="orange", opacity=0.25, line_width=0
                    )

                # Simulating legend for stockout weeks using dummy scatter points
                legend_df = pd.DataFrame({
                    "Working Week": [None, None],  # Invisible points
                    "Inventory": [None, None],
                    "Type": ["Reactive Stockout Weeks", "Proactive Stockout Weeks"]
                })

                fig_legend = px.scatter(
                    legend_df, x="Working Week", y="Inventory", color="Type",
                    color_discrete_map={"Reactive Stockout Weeks": "red", "Proactive Stockout Weeks": "orange"}
                )

                # Add legend traces to the main figure
                for trace in fig_legend.data:
                    fig_inventory.add_trace(trace)

                st.plotly_chart(fig_inventory)

                fig_wos = px.line(
                    wos_df, x="Working Week", y=["Reactive WoS", "Proactive WoS"],
                    title="Weeks of Supply (WoS) Over Time", labels={"value": "WoS", "variable": "Type"}
                )

                fig_wos.update_xaxes(dtick=5)
                fig_wos.update_layout(yaxis_title="Weeks of Supply")

                # Add shaded vertical rectangles for stockout weeks
                for week in representative_stockout_weeks:
                    fig_wos.add_vrect(
                        x0=week - 0.5, x1=week + 0.5,
                        fillcolor="red", opacity=0.25, line_width=0
                    )

                for week in representative_stockout_weeks_proactive:
                    fig_wos.add_vrect(
                        x0=week - 0.5, x1=week + 0.5,
                        fillcolor="orange", opacity=0.25, line_width=0
                    )

                # Simulating legend for stockout weeks using dummy scatter points
                legend_df = pd.DataFrame({
                    "Working Week": [None, None],  # Invisible points
                    "WoS": [None, None],
                    "Type": ["Reactive Stockout Weeks", "Proactive Stockout Weeks"]
                })

                fig_legend = px.scatter(
                    legend_df, x="Working Week", y="WoS", color="Type",
                    color_discrete_map={"Reactive Stockout Weeks": "red", "Proactive Stockout Weeks": "orange"}
                )

                # Add legend traces to the main figure
                for trace in fig_legend.data:
                    fig_wos.add_trace(trace)

                st.plotly_chart(fig_wos)

                fig_consumption = px.line(consumption_df, x='Working Week', y='Consumption', title='Consumption Over Time')
                fig_consumption.update_xaxes(dtick=5)
                fig_consumption.update_layout(yaxis_title='Consumption')

                st.plotly_chart(fig_consumption)

                # Display stockout information
                if representative_stockout_weeks:
                    st.warning(f"Reactive stockout occurred in weeks: {', '.join(map(str, representative_stockout_weeks))}")
                else:
                    st.success("No reactive stockouts occurred.")

                if representative_stockout_weeks_proactive:
                    st.warning(f"Proactive stockout occurred in weeks: {', '.join(map(str, representative_stockout_weeks_proactive))}")
                else:
                    st.success("No proactive stockouts occurred.")

                st.subheader("Weekly Simulation Events after Monte Carlo")
                # for event in representative_weekly_events:
                #     st.markdown(event)

                weekly_table = DES.extract_weekly_table(representative_weekly_events)
                st.write(weekly_table)

                llm_reasoning.explain_inventory_events(representative_weekly_events, reorder_point, lead_time, lead_time_std_dev, consumption_distribution_params, consumption_type,consumption_best_distribution, order_distribution_params, order_quantity_type, order_distribution_best)