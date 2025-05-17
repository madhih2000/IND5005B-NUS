import pandas as pd
import numpy as np
from numpy import ndarray
import streamlit as st
import re

from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

def load_data_consumption(file):
    """
    Loads data from an Excel file and selects specific columns,
    performing necessary data type conversions and cleaning.

    Args:
        file (streamlit.runtime.uploaded_file_manager.UploadedFile):
            The uploaded Excel file.

    Returns:
        pandas.DataFrame: The DataFrame with selected columns and processed data.
    """
    df = pd.read_excel(file)
    # Strip leading and trailing spaces from all string columns
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df.columns = df.columns.str.strip()

    # Select specific columns
    selected_cols = ['Material Group', 'Material Number', 'Pstng Date', 'Quantity',
                     'BUn', 'Plant', 'Site', 'Batch', 'SLED/BBD', 'Vendor Number']
    df = df[selected_cols]

    # Convert 'Pstng Date' to datetime
    df['Pstng Date'] = pd.to_datetime(df['Pstng Date'])

    # Convert 'SLED/BBD' to datetime, handling errors and filling NaT
    df['SLED/BBD'] = pd.to_datetime(df['SLED/BBD'], errors='coerce')
    df['SLED/BBD'] = df['SLED/BBD'].fillna(pd.to_datetime('2100-01-01'))

    # Convert negative consumption values to positive
    df['Quantity'] = df['Quantity'].abs()

    return df

def load_data_consumption_waterfall(file):
    """
    Loads data from an Excel file and selects specific columns,
    performing necessary data type conversions and cleaning.

    Args:
        file (streamlit.runtime.uploaded_file_manager.UploadedFile):
            The uploaded Excel file.

    Returns:
        pandas.DataFrame: The DataFrame with selected columns and processed data.
    """

    df = pd.read_excel(file)
    # Strip leading and trailing spaces from all string columns
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df.columns = df.columns.str.strip()

    # Select specific columns
    selected_cols = ['Material Group', 'Material Number', 'Pstng Date', 'Quantity',
                     'BUn', 'Plant', 'Site', 'Batch', 'SLED/BBD', 'Vendor Number']
    df = df[selected_cols]

    # Convert 'Pstng Date' to datetime
    df['Pstng Date'] = pd.to_datetime(df['Pstng Date'])

    # Add Work Week (WW) column
    iso_calendar = df['Pstng Date'].dt.isocalendar()
    df['WW'] = 'WW' + iso_calendar['week'].astype(str).str.zfill(2)

    # Convert 'SLED/BBD' to datetime, handling errors and filling NaT
    df['SLED/BBD'] = pd.to_datetime(df['SLED/BBD'], errors='coerce')
    df['SLED/BBD'] = df['SLED/BBD'].fillna(pd.to_datetime('2100-01-01'))

    # Convert negative consumption values to positive
    df['Quantity'] = df['Quantity'].abs()

    return df



def load_data_GR(file):
    """
    Loads data from an Excel file and selects specific columns,
    performing necessary data type conversions and cleaning.

    Args:
        file (streamlit.runtime.uploaded_file_manager.UploadedFile):
            The uploaded Excel file.

    Returns:
        pandas.DataFrame: The DataFrame with selected columns and processed data.
    """
    df = pd.read_excel(file)
    # Strip leading and trailing spaces from all string columns
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df.columns = df.columns.str.strip()

    # Convert 'Pstng Date' to datetime
    df['Pstng Date'] = pd.to_datetime(df['Pstng Date'])

    # Convert 'SLED/BBD' to datetime, handling errors and filling NaT
    df['SLED/BBD'] = pd.to_datetime(df['SLED/BBD'], errors='coerce')
    df['SLED/BBD'] = df['SLED/BBD'].fillna(pd.to_datetime('2100-01-01'))

    # Convert negative consumption values to positive
    df['Quantity'] = df['Quantity'].abs()

    return df


def load_data(file):
    df = pd.read_excel(file)
    df['Pstng Date'] = pd.to_datetime(df['Pstng Date'])

    # Strip leading and trailing spaces from all string columns
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df.columns = df.columns.str.strip()

    return df

def load_forecast_consumption_data(file):

    df = pd.read_excel(file)
    df.columns = df.columns.str.strip()

    # Define required columns
    required_columns = {'Material Number', 'Pstng Date', 'Quantity'}
    
    # Check if all required columns are present
    if not required_columns.issubset(df.columns):
        missing = required_columns - set(df.columns)
        raise ValueError(f"Missing columns in the data: {', '.join(missing)}")
    
    
    # Convert posting date to datetime
    df['Pstng Date'] = pd.to_datetime(df['Pstng Date'])
    
    # Extract the ISO week number
    df['Week'] = df['Pstng Date'].dt.isocalendar().week

    # Take absolute of Quantity before grouping
    df['Quantity'] = df['Quantity'].abs()
    
    # Group by Material Number and Week, summing the quantity
    grouped = df.groupby(['Material Number', 'Week'])['Quantity'].sum().reset_index()
    
    # Pivot the table to have weeks as columns
    pivot_df = grouped.pivot(index='Material Number', columns='Week', values='Quantity').fillna(0)
    
    # Rename the week columns to WW1_Consumption, WW2_Consumption, ..., WW52_Consumption
    pivot_df.columns = [f'WW{week}_Consumption' for week in pivot_df.columns]
    
    # Reset index to bring 'Material Number' back as a column
    result_df = pivot_df.reset_index()
    
    return result_df

def merged_order_gr_PO_analysis(df_order: pd.DataFrame, df_GR: pd.DataFrame) -> pd.DataFrame:
    """
    Merges and analyzes Purchase Order (PO) and Goods Receipt (GR) data to compare 
    order quantities with goods received quantities, aggregated by week.

    Args:
        df_order (pd.DataFrame): DataFrame containing Purchase Order details.
        df_GR (pd.DataFrame): DataFrame containing Goods Receipt details.

    Returns:
        pd.DataFrame: A cleaned and aggregated DataFrame with key ordering and GR insights.
    """

    # Strip column names of whitespace
    df_order.columns = df_order.columns.str.strip()
    df_GR.columns = df_GR.columns.str.strip()

    required_order_cols = ['Purchasing Document', 'Vendor Number', 'Material Number', 
                           'Document Date', 'Plant', 'Order Quantity', 
                           'Stockkeeping unit', 'Material Group']
    required_GR_cols = ['Purchasing Document', 'Material Number', 
                        'Pstng Date', 'Site', 'Quantity']

    # Check if required columns exist
    missing_order_cols = [col for col in required_order_cols if col not in df_order.columns]
    missing_GR_cols = [col for col in required_GR_cols if col not in df_GR.columns]

    if missing_order_cols:
        st.warning(f"Missing columns in df_order: {missing_order_cols}")
    if missing_GR_cols:
        st.warning(f"Missing columns in df_GR: {missing_GR_cols}")
    if missing_order_cols or missing_GR_cols:
        st.stop()


    print("Merging PO and GR data...")

    # Merge on 'Purchasing Document' and 'Material Number'
    merged_df = df_order.merge(
        df_GR,
        how='inner',
        on=['Purchasing Document', 'Material Number'],
        suffixes=('_order', '_GR')
    )

    print(merged_df)

    # Strip column names in case of trailing spaces
    merged_df.columns = merged_df.columns.str.strip()

    # Step 2: Select only the needed columns
    cleaned_df = merged_df[[
        'Purchasing Document',
        'Vendor Number_order',
        'Material Number',
        'Document Date',
        'Pstng Date',
        'Plant_order',
        'Site',
        'Supplier',
        'Order Quantity',
        'Quantity',
        'Stockkeeping unit',
        'Material Group_order'
    ]]

    # Step 3: Rename for clarity
    cleaned_df = cleaned_df.rename(columns={
        'Vendor Number_order': 'Vendor Number',
        'Plant_order': 'Plant',
        'Quantity': 'GR Quantity',
        'Material Group_order': 'Material Group',
        'Document Date': 'Order Date',
        'Pstng Date': 'GR Date'
    })

    # Convert dates to datetime if not already
    cleaned_df['Order Date'] = pd.to_datetime(cleaned_df['Order Date'], errors='coerce')
    cleaned_df['GR Date'] = pd.to_datetime(cleaned_df['GR Date'], errors='coerce')

    # Extract ISO calendar week numbers
    cleaned_df['Order WW'] = cleaned_df['Order Date'].dt.isocalendar().week
    cleaned_df['GR WW'] = cleaned_df['GR Date'].dt.isocalendar().week

    # Sort and reset index
    cleaned_df = cleaned_df.sort_values(
        by=['Purchasing Document', 'Material Number', 'Order Date', 'GR Date']
    ).reset_index(drop=True)

    group_cols = ['Purchasing Document', 'Vendor Number', 'Material Number', 'GR Date']

    # Group and aggregate
    df_grouped = cleaned_df.groupby(group_cols, as_index=False).agg({
        'Order Date': 'first',
        'Plant': 'first',
        'Site': 'first',
        'Supplier': 'first',
        'Order Quantity': 'first',
        'GR Quantity': 'sum',
        'Stockkeeping unit': 'first',
        'Material Group': 'first',
        'Order WW': 'first',
        'GR WW': 'first'
    })

    # Final desired column order
    desired_order = [
        'Purchasing Document', 'Vendor Number', 'Material Number',
        'Order Date', 'GR Date', 'Plant', 'Site', 'Supplier',
        'Order Quantity', 'GR Quantity', 'Stockkeeping unit',
        'Material Group', 'Order WW', 'GR WW'
    ]

    return df_grouped[desired_order]

def write_analysis_block(sheet, analysis_text: str, label: str = "Explanation:", merge_cols: int = 8, start_row: int = None):
    """
    Appends a labeled and formatted analysis text block to an existing sheet.

    Args:
        sheet: openpyxl Worksheet object
        analysis_text: The text to be written (can contain line breaks and **bold** markers)
        label: Label before the text (default is 'Explanation:')
        merge_cols: Number of columns to merge for the analysis cell (width of the box)
        start_row: Optional specific row to start writing at
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    import re

    # Insert label
    if start_row is None:
        sheet.append([])
        sheet.append([label])
        start_row = sheet.max_row + 1
    else:
        sheet.cell(row=start_row, column=1, value=label)
        start_row += 1

    # Estimate rows needed
    avg_chars_per_line = 80
    total_chars = len(analysis_text)
    text_lines = analysis_text.count("\n") + total_chars // avg_chars_per_line + 1
    merge_rows = max(5, text_lines)

    end_row = start_row + merge_rows - 1
    start_col = 1
    end_col = merge_cols
    merge_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    sheet.merge_cells(merge_range)

    # Handle bold formatting inside the merged cell
    cell = sheet.cell(row=start_row, column=start_col)
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    if "**" in analysis_text:
        from openpyxl.rich_text import CellRichText, TextBlock
        parts = re.split(r"(\*\*[^*]+\*\*)", analysis_text)
        rich_text = CellRichText()
        for part in parts:
            if part.startswith("**") and part.endswith("**"):
                rich_text.append(TextBlock(part[2:-2], Font(bold=True)))
            else:
                rich_text.append(TextBlock(part))
        cell.value = rich_text
    else:
        cell.value = analysis_text

    # Set row heights
    for r in range(start_row, end_row + 1):
        sheet.row_dimensions[r].height = 20

def sanitize_row(row):
    return [v.item() if isinstance(v, ndarray) and v.size == 1 else v for v in row]