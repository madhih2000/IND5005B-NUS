import plotly.graph_objects as go
import pandas as pd
import numpy as np
import streamlit as st
import os
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict 
from sklearn.preprocessing import MinMaxScaler


def style_dataframe(filtered_df):
    # Identify week columns (WW12, WW13, etc.)
    ww_cols = [col for col in filtered_df.columns if col.startswith('WW')]

    # Define a row-wise styling function
    def highlight_row(row):
        lead_time = row['LeadTime(Week)']
        styles = []
        for col in row.index:
            if col in ww_cols:
                val = row[col]
                if pd.isna(val):
                    styles.append('')
                else:
                    try:
                        val = float(val)
                        if val < 0:
                            styles.append('background-color: #FF5252')  # Red
                        elif val < lead_time:
                            styles.append('background-color: #FFEB3B')  # Yellow
                        else:
                            styles.append('background-color: #4CAF50')  # Green
                    except:
                        styles.append('')
            else:
                styles.append('')
        return styles

    # Apply styling row by row
    styled_df = filtered_df.style.apply(highlight_row, axis=1)

    return styled_df

def generate_weeks_range(start_week, num_weeks=12):
    #weeks_range = [f"WW{str((start_week + i - num_weeks) % 52 or 52).zfill(2)}" for i in range(2 * num_weeks + 1)]
    weeks_range = []
    for i in range(2 * num_weeks + 1):
        week = (start_week + i - num_weeks) % 52 or 52
        if start_week-num_weeks < 1 and week > start_week + num_weeks: #edge case in case you want to check shortage data for week 2 ,e.g.
            continue
        weeks_range.append(f"WW{str(week).zfill(2)}")
    print(weeks_range)
    return weeks_range

# Function to normalize column names
def normalize(col):
    return col.strip().replace('\n', ' ').replace('\r', ' ').replace('\t', ' ').lower()

def save_to_excel(df, output_file):
    """Saves a DataFrame to an Excel file."""
    if df is not None:
        try:
            df.to_excel(output_file, index=False)
            print(f"Data saved to '{output_file}'.")
        except Exception as e:
            print(f"Error saving to Excel: {e}")
    else:
        print("No data to save.")

def extract_and_aggregate_weekly_data(folder_path, material_number, plant, site, start_week,cons_df, num_weeks=12):
    """
    Extracts and aggregates weekly data for a specific material number, plant, and site,
    starting from a specified week and including the next 'num_weeks' weeks.

    Args:
        folder_path (str): Path to the folder containing the XLSX files.
        material_number (str): Material number to filter.
        plant (str): Plant to filter.
        site (str): Site to filter.
        start_week (str): Starting week (e.g., "WW32").
        num_weeks (int): Number of weeks to include in the output.

    Returns:
        pandas.DataFrame: DataFrame containing the aggregated weekly data, or None if no data is found.
    """

    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' not found.")
        return None
    
    print(folder_path)

    all_files = sorted([f for f in os.listdir(folder_path) if f.endswith(".xlsx")])

    if not all_files:
        print(f"Error: No XLSX files found in '{folder_path}'.")
        return None

    # Filter the Consumption DataFrame based on the selected values
    cons_df_filtered = cons_df[
        (cons_df["Material Number"] == material_number) &
        (cons_df["Plant"] == plant) &
        (cons_df["Site"] == site)
    ]

    if cons_df_filtered.empty:
        st.warning("Consumption data does not exist for the selected material number, plant, and site.")

    # Aggregate consumption data by WW
    cons_agg = (
        cons_df_filtered[
            (cons_df["Material Number"] == material_number) &
            (cons_df["Plant"] == plant) &
            (cons_df["Site"] == site)
        ]
        .groupby("WW")["Quantity"]
        .sum()
        .reset_index()
    )

    # Normalize column names for consistency
    cons_agg["WW"] = cons_agg["WW"].str.upper()

    week_numbers = list(range(max(1, start_week - num_weeks), start_week + 1))
    #print(week_numbers)

    selected_data = []
    selected_weeks = []

    #Columns to add to the final df
    initial_columns = ["MaterialNumber", "Plant", "Site", "Measures", "InventoryOn-Hand", "LeadTime(Week)"]
    weeks_range = generate_weeks_range(start_week,num_weeks)
    #print(weeks_range)

    for i in week_numbers:
        week_file = f"WW{i}.xlsx"
        week_col = f"WW{i:02d}"

        if week_file in all_files:
            file_path = os.path.join(folder_path, week_file)
            try:
                df = pd.read_excel(file_path)
                filtered_df = df[
                    (df["Material Number"] == material_number)
                    & (df["Plant"] == plant)
                    & (df["Site"] == site)
                ]

                lead_value = 0

                if i == start_week:

                    # Lead Time column, normalized
                    target_col = normalize("Lead Time (Week)")
                    lead_col = None

                    # Loop to find the unnormalized matching column
                    for col in df.columns:
                        if normalize(col) == target_col:
                            lead_col = col
                            break

                    if lead_col:
                        lead_value = df.loc[df["Material Number"] == material_number, lead_col]

                
                #Remove whitespace in column names
                filtered_df.columns = filtered_df.columns.str.strip().str.replace('\n', ' ', regex=False).str.replace(r'\s+', '', regex=True)
        
                if not filtered_df.empty:
                    # Select only the required columns
                    all_columns = initial_columns + weeks_range

                    # Check if all columns exist in filtered_df
                    missing_columns = [col for col in all_columns if col not in filtered_df.columns]
                    
                    if missing_columns:
                        raise ValueError(f"Missing columns in {week_file}: {', '.join(missing_columns)}")
                        return None
                    
                    filtered_df = filtered_df[all_columns]

                    #Add new column snapshot
                    filtered_df['Snapshot'] = week_col

                    # Remove the first element from weeks_range for next iteration
                    weeks_range = weeks_range[1:]

                    selected_data.append(filtered_df)
                    selected_weeks.append(f"WW{i:02d}")

                else:
                    # If filtered_df is empty, create a new dataframe with NaN values
                    all_columns = initial_columns + weeks_range
                    empty_df = pd.DataFrame(columns=all_columns)
                    empty_df['Snapshot'] = week_col  # Add the Snapshot column with the week_col value

                    # Add the empty dataframe to selected_data and weeks
                    selected_data.append(empty_df)
                    selected_weeks.append(f"WW{i:02d}")

            except FileNotFoundError:
                print(f"Error: File '{week_file}' not found.")
            except Exception as e:
                print(f"Error reading file '{week_file}': {e}")
        else:
            print(f"Warning: Week file '{week_file}' not found.")

    if not selected_data:
        print("No matching data found.")
        return None

    result_df = pd.concat(selected_data, ignore_index=True)
    
    #reorder columns to have week at the beginning.
    cols = result_df.columns.tolist()
    cols = ['Snapshot'] + [col for col in cols if col != 'Snapshot']
    result_df = result_df[cols]

    result_df = adding_consumption_data_from_agg(result_df, cons_agg)

    return result_df, lead_value

def adding_consumption_data_from_agg(result_df, cons_agg):
    """
    Modifies result_df in-place by filling 'Consumption' rows using cons_agg.
    Adds a Consumption row for every week in result_df.
    If a quantity exists in cons_agg for that week, use that quantity, otherwise use 0.
    Each value is added on the diagonal (Snapshot == WW column), and
    other week columns are set to None.

    Args:
        result_df (pd.DataFrame): The target DataFrame with Snapshot and week columns.
        cons_agg (pd.DataFrame): DataFrame with 'WW' and 'Quantity' columns.

    Returns:
        pd.DataFrame: Updated DataFrame with Consumption values applied.
    """
    df = result_df.copy()

    # Get week columns from the unique values of the 'Snapshot' column.
    week_cols = df['Snapshot'].unique().tolist()
    week_cols = [col for col in week_cols if str(col).startswith("WW")] # added str()

    # Get example metadata row (first Supply or Demand row)
    meta_row = df[df['Measures'].isin(['Supply', 'Alternate Demand'])].iloc[0]

    # Store new rows
    new_rows = []

    # Create a set of weeks from cons_agg for faster lookup
    cons_agg_weeks = set(cons_agg['WW'])

    for week in week_cols:
        # Create base row
        new_row = meta_row.copy()
        new_row['Measures'] = 'Consumption'
        new_row['Snapshot'] = week
        new_row['InventoryOn-Hand'] = None
        
        
        # Get the LeadTime(Week) for the current snapshot
        lead_time_row = df[df['Snapshot'] == week].iloc[0] # Get the first row for the snapshot
        new_row['LeadTime(Week)'] = lead_time_row['LeadTime(Week)']

        # Set all week columns to None
        for col in week_cols:
            new_row[col] = None

        # Set quantity in diagonal cell
        if week in cons_agg_weeks:
            quantity = cons_agg[cons_agg['WW'] == week]['Quantity'].values[0]  # Get quantity from cons_agg
            new_row[week] = quantity
        else:
            new_row[week] = 0  # Default quantity if week not in cons_agg

        new_rows.append(new_row)

    # Append all new consumption rows
    consumption_df = pd.DataFrame(new_rows)
    df = pd.concat([df, consumption_df], ignore_index=True)

    # Optional: sort to keep consistent order.  Important for comparisons.
    df.sort_values(by=['Snapshot', 'Measures'], inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def plot_stock_prediction_plotly(df, start_week, lead_time, weeks_range):
    """Plots actual and predicted stock values over weeks."""

    weeks = generate_weeks_range(start_week, weeks_range)
    stock_values = df[df['Measures'] == 'Weeks of Stock']

    fig = go.Figure()

    actual_values = []
    predicted_values = []
    plot_weeks_actual = []
    plot_weeks_predicted = []

    for i, week in enumerate(weeks):
        print(i)
        print(predicted_values)
        if week not in df.columns:
            print(f"Warning: Week {week} not found in DataFrame.")
            continue  # skip the iteration if the week is not in the dataframe.

        if week in df['Snapshot'].values: # Only get actual values if the week is in the snapshot.
            actual_value = stock_values[stock_values['Snapshot'] == week][week].iloc[0]
            if actual_value is None or np.isnan(actual_value):
                actual_value = 0  # if actual value is none or nan, set to 0.
            actual_values.append(actual_value)
            plot_weeks_actual.append(week)
        else:
            break

        if i != 0:
            predicted_value = stock_values[stock_values['Snapshot'] == weeks[i - 1]][week].iloc[0]
        else:
            predicted_value = None

        if predicted_value is None or np.isnan(predicted_value):
            if i > 0:
                predicted_value = predicted_values[-1]  # use the previous predicted value if the current one is None or NaN
            else:
                predicted_value = 0  # skip the first one as it is none.

        predicted_values.append(predicted_value)
        plot_weeks_predicted.append(week)

    # Create comparison DataFrame
    comparison = pd.DataFrame({
        'Week': plot_weeks_actual,
        'Actual': actual_values,
        'Predicted': predicted_values[:len(plot_weeks_actual)]  # Align lengths
    })

    # Add Deviation Flag
    def check_deviation_detail(row):
        if row['Actual'] == 0 and row['Predicted'] == 0:
            return False, ""
        elif row['Actual'] == 0 and row['Predicted'] != 0:
            return True, "Actual demand is 0, but forecasted is not"
        elif row['Actual'] != 0 and row['Predicted'] == 0:
            return True, "Forecasted demand expected to be 0, but actual is not"
        ratio = row['Predicted'] / row['Actual']
        if ratio >= 1.5:
            percent = round((ratio - 1) * 100, 1)
            return True, f"Overestimate by {percent}%"
        elif ratio <= (1 / 1.5):
            percent = round((1 - ratio) * 100, 1)
            return True, f"Underestimate by {percent}%"
        else:
            return False, ""

    deviation_results = comparison.apply(check_deviation_detail, axis=1, result_type='expand')
    comparison['Deviation_Flag'] = deviation_results[0]
    comparison['Deviation_Detail'] = deviation_results[1]

    fig.add_trace(go.Scatter(x=plot_weeks_actual, y=actual_values, mode='lines+markers', name='Actual Weeks of Stock'))
    fig.add_trace(go.Scatter(x=plot_weeks_predicted, y=predicted_values, mode='lines+markers', name='Predicted Weeks of Stock'))

    fig.update_layout(title='Actual vs. Predicted Weeks of Stock',
                      xaxis_title='Week',
                      yaxis_title='Weeks of Stock')

    return actual_values, fig, comparison

def plot_consumption_vs_demand_plotly(waterfall_df):
    """Plots and compares Consumption vs Demand w/o Buffer from a Waterfall DataFrame."""

    # Extract required rows
    demand_df = waterfall_df[waterfall_df['Measures'] == 'Demand w/o Buffer']
    consumption_df = waterfall_df[waterfall_df['Measures'] == 'Consumption']

    # Weeks to evaluate
    snapshots = list(waterfall_df['Snapshot'].unique())
    common_weeks = [w for w in snapshots if w in demand_df.columns and w in consumption_df.columns]

    actual_demand = []
    actual_consumption = []
    plot_weeks = []

    for week in common_weeks:
        demand_row = demand_df[demand_df['Snapshot'] == week]
        cons_row = consumption_df[consumption_df['Snapshot'] == week]

        if demand_row.empty or cons_row.empty:
            continue

        demand = demand_row[week].iloc[0]
        consumption = cons_row[week].iloc[0]

        # Handle missing or NaN values
        demand = 0 if pd.isna(demand) else demand
        consumption = 0 if pd.isna(consumption) else consumption

        actual_demand.append(demand)
        actual_consumption.append(consumption)
        plot_weeks.append(week)

    # Build DataFrame for analysis
    comparison = pd.DataFrame({
        'Week': plot_weeks,
        'Demand': actual_demand,
        'Consumption': actual_consumption,
    })

    # Deviation logic
    def analyze_discrepancy(row):
        if row['Demand'] == 0 and row['Consumption'] == 0:
            return False, ""
        elif row['Demand'] == 0 and row['Consumption'] != 0:
            return True, "Demand is 0 but consumption occurred"
        elif row['Demand'] != 0 and row['Consumption'] == 0:
            return True, "Demand expected but no consumption"
        ratio = row['Consumption'] / row['Demand']
        if ratio >= 1.5:
            return True, f"Over-consumption by {round((ratio - 1) * 100, 1)}%"
        elif ratio <= (1 / 1.5):
            return True, f"Under-consumption by {round((1 - ratio) * 100, 1)}%"
        else:
            return False, ""

    flags = comparison.apply(analyze_discrepancy, axis=1, result_type='expand')
    comparison['Deviation_Flag'] = flags[0]
    comparison['Deviation_Detail'] = flags[1]

    # Plotting
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=plot_weeks, y=actual_demand, mode='lines+markers', name='Demand'))
    fig.add_trace(go.Scatter(x=plot_weeks, y=actual_consumption, mode='lines+markers', name='Consumption'))

    fig.update_layout(
        title='Consumption vs Demand (Waterfall)',
        xaxis_title='Week',
        yaxis_title='Quantity',
        legend=dict(x=0.01, y=0.99)
    )

    return actual_consumption, fig, comparison

def identify_specific_po_timing_issues(demand_df, po_df):
    """
    Scenario 2: Match POs to actual weekly demand using Inventory-On-Hand
    and provide actionable push/pull suggestions.

    Args:
        demand_df: DataFrame containing waterfall data with 'Snapshot', 'Measures', and WW columns.
        po_df: DataFrame with ['Purchasing Document', 'Order WW', 'GR WW', 'GR Quantity'].

    Returns:
        A DataFrame showing demand coverage, PO matching, unmet demand, and suggested PO timing actions.
    """
    demand_rows = demand_df[demand_df['Measures'] == 'Demand w/o Buffer']
    supply_rows = demand_df[demand_df['Measures'] == 'Supply']  # corrected to use Supply for inventory
    weeks = sorted([col for col in demand_df.columns if col.startswith("WW")])
    snapshots = demand_rows['Snapshot'].unique()

    po_df = po_df.copy()
    po_df['Remaining Qty'] = po_df['GR Quantity']

    results = []

    for snapshot in snapshots:
        week = snapshot  # e.g., WW08
        week_num = int(week.replace("WW", ""))

        # Get demand for this actual week
        demand_val = demand_rows[demand_rows['Snapshot'] == snapshot][week]
        demand = int(demand_val.values[0]) if not demand_val.empty and not pd.isna(demand_val.values[0]) else 0

        # ✅ Corrected: Get Inventory-On-Hand from the 'Supply' row using the week column
        inv_val = supply_rows[supply_rows['Snapshot'] == snapshot]['InventoryOn-Hand']
        inventory_on_hand = int(inv_val.values[0]) if not inv_val.empty and not pd.isna(inv_val.values[0]) else 0

        matched_pos = []
        remaining_demand = demand

        # Use inventory first
        inventory_used = min(inventory_on_hand, remaining_demand)
        remaining_demand -= inventory_used

        # Use POs arriving in same week
        available_pos = po_df[(po_df['GR WW'] == week_num) & (po_df['Remaining Qty'] > 0)]
        for _, po in available_pos.iterrows():
            po_id = po['Purchasing Document']
            qty = po['Remaining Qty']
            used = min(qty, remaining_demand)

            po_df.loc[po_df['Purchasing Document'] == po_id, 'Remaining Qty'] -= used
            matched_pos.append(f"{po_id} (Used {used})")
            remaining_demand -= used

            if remaining_demand <= 0:
                break

        # Determine if POs need adjustment
        flags = []

        if remaining_demand > 0:
            # Pull in future POs
            late_pos = po_df[(po_df['GR WW'] > week_num) & (po_df['Remaining Qty'] > 0)]
            for _, po in late_pos.iterrows():
                flags.append(f"Pull In: PO {po['Purchasing Document']} (ETA WW{po['GR WW']}, Qty {po['Remaining Qty']})")

        # Check if there are early POs not needed yet
        early_pos = po_df[(po_df['GR WW'] < week_num) & (po_df['Remaining Qty'] > 0)]
        for _, po in early_pos.iterrows():
            flags.append(f"Push Out: PO {po['Purchasing Document']} (ETA WW{po['GR WW']}, Qty {po['Remaining Qty']})")

        results.append({
            "Snapshot Week": snapshot,
            "Week": week,
            "Demand (Actual)": demand,
            "Inventory-On-Hand": inventory_on_hand,
            "Inventory Used": inventory_used,
            "Matched PO Lines": ", ".join(matched_pos) if matched_pos else "None",
            "Unmet Demand": max(0, remaining_demand),
            "Actions Needed": "; ".join(flags) if flags else "OK"
        })

    return pd.DataFrame(results)

def check_wos_against_lead_time(wos_list, lead_time):
    """
    Compares values in wos_list against lead_time and flags close matches.

    Args:
        wos_list: A list of floats, potentially containing 0.
        lead_time: An integer representing the target lead time.

    Returns:
        A list of messages indicating the comparison results and a boolean indicating
        whether an immediate order is needed.
    """
    # Check if lead_time is a pandas Series and handle it
    if isinstance(lead_time, pd.Series):
        lead_time_value = lead_time.iloc[0]
    else:
        lead_time_value = lead_time  # If it's already a scalar


    messages = []
    order_immediately = False
    tolerance = 0.1  # 10% tolerance

    for i, wos in enumerate(wos_list):
        if wos == 0:
            #messages.append(f"Index {i}: No value recorded.")
            continue

        lower_bound = lead_time_value * (1 - tolerance)
        upper_bound = lead_time_value * (1 + tolerance)

        if wos < lower_bound:
            messages.append(f"Index {i}: Value {wos} is significantly lower than lead time {lead_time_value}. ALERT!")
        elif wos > upper_bound:
            continue
            #messages.append(f"Index {i}: Value {wos} is slightly higher than lead time {lead_time_value}. Warning.")

    if wos_list and wos_list[-1] != 0:
        if wos_list[-1] < lead_time_value*0.5: # if the last value is less than half the lead time.
            order_immediately = True
            messages.append("Last recorded value is significantly low, consider immediate order.")
        else:
            messages.append("Last recorded value does not indicate immediate order is needed.")

    return messages, order_immediately

def apply_coloring_to_output(excel_buffer, lead_time, sheet_names):
    # Rewind buffer and load workbook
    # Rewind buffer and load workbook
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)

    # Fill colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Detect header row
        header_row_idx = 2 if sheet_name == "RCA Scenario 1" else 1

        header = [cell.value for cell in next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))]
        ww_col_indices = [i + 1 for i, h in enumerate(header) if str(h).startswith("WW")]

        if not ww_col_indices:
            continue

        if sheet_name == "Waterfall Chart":
            try:
                measures_col_idx = header.index("Measures") + 1
            except ValueError:
                continue

            for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
                if row[measures_col_idx - 1].value == 'Weeks of Stock':
                    for idx in ww_col_indices:
                        cell = row[idx - 1]
                        try:
                            val = float(cell.value)
                            if val < 0:
                                cell.fill = red_fill
                            elif val < lead_time.iloc[0]:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = green_fill
                        except (TypeError, ValueError):
                            continue

        elif sheet_name == "RCA Scenario 1":
            for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
                for idx in ww_col_indices:
                    cell = row[idx - 1]
                    try:
                        val = float(cell.value)
                        if val < 0:
                            cell.fill = red_fill
                        elif val < lead_time.iloc[0]:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = green_fill
                    except (TypeError, ValueError):
                        continue

    colored_output = BytesIO()
    wb.save(colored_output)
    colored_output.seek(0)
    return colored_output

#Condition 4: Longer delivery lead time
def lead_time_check(result_df):
    # First, group by Snapshot and collect the unique LeadTime(Week) values per snapshot
    leadtime_changes = ( 
    result_df.groupby("Snapshot")["LeadTime(Week)"]
    .unique()
    .reset_index()
    )

    # function to check for changes and generate a string of changes
    def detect_changes(row):
        leadtimes = row["LeadTime(Week)"]
        if len(leadtimes) > 1:
            change_str = f"Lead time changes detected: {', '.join(map(str, leadtimes))}"
            return True, change_str
        else:
            return False, "No change"

    # Apply the function to create new columns
    leadtime_changes[["Changed", "Change Details"]] = leadtime_changes.apply(
        detect_changes, axis=1, result_type="expand"
    )

    return leadtime_changes

# Scenario 5: Identifying irregular consumption patterns
def analyze_week_to_week_demand_changes(result_df, abs_threshold=10, pct_threshold=0.3, lead_time = 6):
    """
    Filters for 'Demand w/o Buffer', calculates demand values based on snapshot columns,
    and detects week-to-week anomalies such as spikes or drops.

    Returns:
        - output_df: DataFrame with demand values and irregular pattern flags
    """
    # Step 1: Filter
    filtered_df = result_df[result_df['Measures'] == 'Demand w/o Buffer'].copy()
    if filtered_df.empty:
        raise ValueError("No rows found with Measures == 'Demand w/o Buffer'")

    # Find all WW columns
    ww_cols = [col for col in filtered_df.columns if col.startswith("WW")]

    output_all = []

    lead_time = lead_time.iloc[0] #lead time taking in as a series

    for i in range(1, len(ww_cols)):  # start from 2nd WW column
        current_col = ww_cols[i]

        # Determine the window size
        window_size = i + 1

        # Determine row indices
        if i < (lead_time+1):
            window_data = filtered_df[current_col].iloc[:window_size]
            snapshot_ids = filtered_df.index[:window_size]
        else:
            window_data = filtered_df[current_col].iloc[-(lead_time+1):]
            snapshot_ids = filtered_df.index[-(lead_time+1):]

        # Build temporary DataFrame
        output_df = pd.DataFrame({
            "Snapshot": snapshot_ids,
            "Demand w/o Buffer": window_data.values,
            "Week": current_col
        })

        # Sort and calculate WoW change
        output_df = output_df.sort_values(by="Snapshot").reset_index(drop=True)
        output_df["WoW Change"] = output_df["Demand w/o Buffer"].diff()
        output_df["WoW % Change"] = output_df["Demand w/o Buffer"].pct_change()

        # Flag spikes and drops
        output_df["Spike"] = output_df["WoW Change"] > abs_threshold
        output_df["Drop"] = output_df["WoW Change"] < -abs_threshold
        output_df["Sudden % Spike"] = output_df["WoW % Change"] > pct_threshold
        output_df["Sudden % Drop"] = output_df["WoW % Change"] < -pct_threshold

        # Store for analysis
        output_all.append(output_df)

    # Combine all weeks’ flagged data
    final_output_df = pd.concat(output_all, ignore_index=True)
    final_output_df = final_output_df.drop(columns=['Snapshot'])
    cols = final_output_df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('Week')))
    final_output_df = final_output_df[cols]

    return final_output_df


def calculate_weekly_demand_summary(data):
    """
    Returns a simplified weekly demand variability summary,
    including SD, anomaly counts, average WoW changes, and an irregularity score.
    """
    # Standard Deviation (SD)
    sd = data.groupby("Week")["Demand w/o Buffer"].std().reset_index()
    sd.columns = ["Week", "SD"]
    sd["SD"] = sd["SD"].round(2)

    # Anomaly Counts
    anomaly_counts = data.groupby("Week")[["Spike", "Drop", "Sudden % Spike", "Sudden % Drop"]].sum().reset_index()

    # Avg Absolute WoW Change
    data["Abs WoW Change"] = data["WoW Change"].abs()
    avg_abs_change = data.groupby("Week")["Abs WoW Change"].mean().reset_index()
    avg_abs_change.columns = ["Week", "Avg Abs WoW Change"]
    avg_abs_change["Avg Abs WoW Change"] = avg_abs_change["Avg Abs WoW Change"].round(2)

    # Merge all tables in order of 'Week' appearance in original data
    summary = sd.merge(anomaly_counts, on="Week").merge(avg_abs_change, on="Week")

    # Preserve original Week order from the input data
    week_order = data["Week"].drop_duplicates().tolist()
    summary["Week"] = pd.Categorical(summary["Week"], categories=week_order, ordered=True)
    summary = summary.sort_values("Week").reset_index(drop=True)

    # Irregularity Score 
    scaler = MinMaxScaler()
    score_cols = ["SD", "Spike", "Drop", "Avg Abs WoW Change"]
    summary["Irregularity Score"] = scaler.fit_transform(summary[score_cols]).sum(axis=1).round(2)

    return summary

def scenario_6(waterfall_df, po_df):
    # Filter relevant rows
    supply_rows = waterfall_df[waterfall_df['Measures'] == 'Supply']
    demand_rows = waterfall_df[waterfall_df['Measures'] == 'Demand w/o Buffer']
    consumption_rows = waterfall_df[waterfall_df['Measures'] == 'Consumption']

    # Get initial snapshot and inventory from Supply rows
    initial_snapshot = supply_rows['Snapshot'].iloc[0]
    initial_inventory_calc = int(supply_rows[supply_rows['Snapshot'] == initial_snapshot]['InventoryOn-Hand'].values[0])

    # All snapshots to iterate through
    snapshots = list(waterfall_df['Snapshot'].unique())

    results = []
    current_inventory_calc = initial_inventory_calc

    for i, snapshot in enumerate(snapshots):
        week_col = snapshot
        week_num = int(snapshot.replace("WW", ""))

        # Get supply and demand values
        demand_val = demand_rows[demand_rows['Snapshot'] == snapshot][week_col]
        supply_val = supply_rows[supply_rows['Snapshot'] == snapshot][week_col]

        demand = int(demand_val.values[0]) if not demand_val.empty else 0
        supply = int(supply_val.values[0]) if not supply_val.empty else 0

        # Get consumption value
        consumption_val = consumption_rows[consumption_rows['Snapshot'] == snapshot][week_col]
        consumption_waterfall = int(consumption_val.values[0]) if not consumption_val.empty else 0

        # Start inventory from Waterfall column
        inv_val = supply_rows[supply_rows['Snapshot'] == snapshot]['InventoryOn-Hand']
        start_inventory_waterfall = int(inv_val.values[0]) if not inv_val.empty else 0

        # GR quantity from PO data
        po_received = 0
        if not po_df.empty:
            po_received = po_df[po_df['GR WW'] == week_num]['GR Quantity'].sum()

        # Calculate end inventories
        end_inventory_calc = current_inventory_calc + supply - consumption_waterfall
        end_inventory_waterfall = start_inventory_waterfall + supply - consumption_waterfall

        # Calculate consumption (Calc) using NEXT week's inventory
        if i < len(snapshots) - 1:
            next_snapshot = snapshots[i + 1]
            next_inv_val = supply_rows[supply_rows['Snapshot'] == next_snapshot]['InventoryOn-Hand']
            next_start_inventory = int(next_inv_val.values[0]) if not next_inv_val.empty else 0
            consumption_calc = next_start_inventory - (start_inventory_waterfall + supply)
        else:
            consumption_calc = 0  # No next week to compare

                # Check for irregular patterns
        consumption_diff = consumption_calc - consumption_waterfall
        irregular_pattern = []

        if consumption_waterfall < 0:
            if results:
                previous_week = results[-1]
                prev_end_inv = previous_week['End Inventory (Waterfall)']
                prev_demand = previous_week['Demand (Waterfall)']
                prev_consumption = previous_week['Consumption (Waterfall)']
                prev_supply = previous_week['Supply (Waterfall)']
                prev_snapshot = previous_week['Snapshot Week']

                irregular_pattern.append(
                    f"In week {snapshot}, the reported consumption was negative, which suggests a possible inventory correction or data anomaly. "
                    f"Looking back, in week {prev_snapshot}, the starting inventory was {previous_week['Start Inventory (Waterfall)']}, "
                    f"supply was {prev_supply}, and demand was {prev_demand}, resulting in an expected end inventory of {prev_end_inv}. "
                    f"However, the current week's starting inventory ({start_inventory_waterfall}) is higher than the expected ending inventory from last week, "
                    f"which implies inventory was added back — potentially due to returns, cancellations, or adjustments outside the normal flow."
                )
            else:
                irregular_pattern.append(
                    f"In week {snapshot}, negative consumption was reported. Since this is the first week in the timeline, "
                    f"it may indicate an opening adjustment or return to inventory that wasn't accounted for in demand."
                )
        elif consumption_waterfall > demand:
            irregular_pattern.append("More consumption than demand")
        elif consumption_waterfall == 0 and demand != 0:
            irregular_pattern.append("Consumption is zero but demand is not")
        elif consumption_waterfall != 0 and demand == 0:
            irregular_pattern.append("Demand is zero but consumption is not")

        # Flag and explain if calculated and reported consumption differ significantly
        if abs(consumption_diff) > 5 and i < len(snapshots) - 1:
            if consumption_diff > 0:
                reasoning = (
                    "This suggests that more inventory was consumed (based on stock movement) than what was reported. "
                    "Possible causes include unrecorded consumption events, scrap/write-offs, or timing mismatches where stock depletion occurred but wasn't captured in the consumption metric."
                )
            else:
                reasoning = (
                    "This suggests that reported consumption was higher than what is reflected in actual inventory movement. "
                    "Possible causes include unposted supply receipts, early booking of consumption, or inventory adjustments that didn’t flow through standard consumption channels."
                )

            irregular_pattern.append(
                f"In week {snapshot}, the calculated consumption is {consumption_calc}, while the reported consumption is {consumption_waterfall} "
                f"(difference of {consumption_diff}). The calculation used the formula: "
                f"Consumption (Calc) = Next Start Inventory ({next_start_inventory}) − (Current Start Inventory ({start_inventory_waterfall}) + Supply ({supply})). "
                f"{reasoning}"
            )

        # Combine messages
        irregular_pattern_str = " | ".join(irregular_pattern) if irregular_pattern else None

        results.append({
            'Snapshot Week': snapshot,
            'Start Inventory (Waterfall)': start_inventory_waterfall,
            'Start Inventory (Calc)': current_inventory_calc,
            'Demand (Waterfall)': demand,
            'Supply (Waterfall)': supply,
            'Consumption (Waterfall)': consumption_waterfall,
            'Consumption (Calc)': consumption_calc,
            'PO GR Quantity': po_received,
            'End Inventory (Waterfall)': end_inventory_waterfall,
            'End Inventory (Calc)': end_inventory_calc,
            'Irregular Pattern': irregular_pattern_str
        })

        current_inventory_calc = end_inventory_calc

    return pd.DataFrame(results)

def scenario_1(df, po_df):
    # Filter for 'Weeks of Stock' rows
    weeks_df = df[df['Measures'] == 'Weeks of Stock'].copy()
    weeks_df = weeks_df.reset_index(drop=True)

    # Get all WW columns and ensure they're ordered
    week_cols = sorted([col for col in df.columns if col.startswith('WW')])

    # Function to get week range per row
    def get_leadtime_cols(snapshot, leadtime, all_weeks):
        if snapshot not in all_weeks:
            return []
        start_idx = all_weeks.index(snapshot)
        end_idx = start_idx + leadtime  # snapshot + (leadtime - 1)
        return all_weeks[start_idx:end_idx]

    # Create filtered output
    filtered_rows = []
    for _, row in weeks_df.iterrows():
        leadtime = int(row['LeadTime(Week)'])
        snapshot = row['Snapshot']
        cols_to_keep = get_leadtime_cols(snapshot, leadtime, week_cols)

        base_info = row.drop(week_cols + ['InventoryOn-Hand'], errors='ignore')  # Keep non-WW fields
        week_data = row[cols_to_keep]   # Select WW columns in range
        combined = pd.concat([base_info, week_data])
        filtered_rows.append(combined)

    # Final filtered DataFrame
    filtered_df = pd.DataFrame(filtered_rows)

    # Ensure Snapshot and GR WW are numeric and comparable
    po_df['GR WW'] = po_df['GR WW'].astype(int)
    filtered_df['Snapshot'] = filtered_df['Snapshot'].str.extract(r'(\d+)').astype(int)

    # Function to filter POs based on lead time
    def filter_pos_by_leadtime(row, leadtime, po_df):
        snapshot = row['Snapshot']
        start_week = snapshot
        end_week = snapshot + leadtime - 1
        gr_weeks = list(range(start_week, end_week + 1))
        filtered_po_df = po_df[po_df['GR WW'].isin(gr_weeks)]
        incoming_po = ', '.join(filtered_po_df['Purchasing Document'].astype(str).unique())
        return incoming_po

    # Add Incoming PO column
    filtered_df['Incoming PO'] = filtered_df.apply(
        lambda row: filter_pos_by_leadtime(row, row['LeadTime(Week)'], po_df), axis=1
    )

    def flag_row_with_reason(row):
        leadtime = row['LeadTime(Week)']
        has_incoming_po = row['Incoming PO'] != ''
        week_cols_in_row = [col for col in week_cols if col in row.index]
        values_series = row[week_cols_in_row]
        numeric_values = values_series[values_series.notna()].astype(float)

        if numeric_values.empty:
            return 'Unknown Case - All Weeks Null', 'All Weeks of Stock values are missing or null.'

        below_lt = numeric_values < leadtime
        negative = numeric_values < 0
        adequate = (numeric_values >= leadtime) & (numeric_values >= 0)

        if below_lt.all() or negative.all():
            if has_incoming_po:
                return 'Adequate', 'Stock is low but incoming PO exists within lead time.'
            else:
                return 'Inadequate', 'All WoS values are below lead time and no PO is expected.'
        elif adequate.all():
            return 'Not Applicable', 'All WoS values meet or exceed lead time—no issue.'
        elif below_lt.any() or negative.any():
            if has_incoming_po:
                return 'Partially Adequate', 'Some WoS values are low, but there is an incoming PO.'
            else:
                return 'Partially Inadequate', 'Some WoS values are low and there is no incoming PO.'
        else:
            return 'Unknown Case - Numeric Checks Failed', 'Unexpected data condition—please review values.'

    # Apply and separate into two columns
    filtered_df[['Flag', 'Reason']] = filtered_df.apply(
        flag_row_with_reason, axis=1, result_type='expand'
    )

    return filtered_df

def scenario_3(waterfall_df, po_df, lead_time: int): # Added lead_time argument
    """
    Analyzes inventory based on waterfall data and identifies potential
    pull-in or push-out actions for POs from po_df, based on specific metric conditions.
    This version does not simulate the impact of actions on inventory.

    Args:
        waterfall_df (pd.DataFrame): DataFrame with supply, demand, and inventory snapshots.
                                     Expected columns: 'Measures', 'Snapshot', 'InventoryOn-Hand',
                                     and week columns (e.g., 'WW23').
        po_df (pd.DataFrame): DataFrame with purchase order details.
                              Expected columns: 'Purchasing Document', 'GR WW', 'Order WW', 'GR Quantity'.
        lead_time (int): Lead time in weeks, used for pull-in condition.
    Returns:
        pd.DataFrame: DataFrame with weekly analysis, including suggested actions.
    """
    lead_time = lead_time.iloc[0]
    supply_rows = waterfall_df[waterfall_df['Measures'] == 'Supply']
    demand_rows = waterfall_df[waterfall_df['Measures'] == 'Demand w/o Buffer']

    # Ensure 'Snapshot' is sorted to process weeks chronologically
    snapshots = sorted(waterfall_df['Snapshot'].unique())
    
    if not snapshots:
        print("Warning: No snapshots found in waterfall_df.")
        return pd.DataFrame()

    # Determine initial inventory from the first snapshot
    initial_snapshot_data = supply_rows[supply_rows['Snapshot'] == snapshots[0]]
    if not initial_snapshot_data.empty and 'InventoryOn-Hand' in initial_snapshot_data.columns:
        initial_inventory_calc = int(initial_snapshot_data['InventoryOn-Hand'].iloc[0])
    else:
        print(f"Warning: Could not determine initial inventory from snapshot {snapshots[0]}. Defaulting to 0.")
        initial_inventory_calc = 0

    results = []
    current_inventory_calc = initial_inventory_calc
    # Tracks POs that have been suggested for an action to avoid re-suggesting them
    suggested_action_po_docs_globally = set() 

    for i, snapshot in enumerate(snapshots):
        week_col = snapshot
        try:
            week_num = int(snapshot.replace("WW", ""))
        except ValueError:
            print(f"Warning: Could not parse week number from snapshot {snapshot}. Skipping this snapshot.")
            continue

        # --- Demand & Supply from Waterfall DF for the current week column ---
        demand_val_series = demand_rows.loc[demand_rows['Snapshot'] == snapshot, week_col]
        supply_val_series = supply_rows.loc[supply_rows['Snapshot'] == snapshot, week_col]
        demand_waterfall = int(demand_val_series.iloc[0]) if not demand_val_series.empty else 0
        supply_waterfall = int(supply_val_series.iloc[0]) if not supply_val_series.empty else 0
        
        # --- Waterfall Inventory Reference (Start of Week from 'InventoryOn-Hand' column) ---
        inv_on_hand_series = supply_rows.loc[supply_rows['Snapshot'] == snapshot, 'InventoryOn-Hand']
        start_inventory_waterfall_ref = int(inv_on_hand_series.iloc[0]) if not inv_on_hand_series.empty else 0

        # --- Store current week's opening inventory (carried from previous week's close) ---
        opening_inventory_calc = current_inventory_calc

        # --- Calculated End Inventory (based purely on waterfall_df data for the week) ---
        end_inventory_calc = opening_inventory_calc + supply_waterfall - demand_waterfall
        
        action = "No Action"
        suggested_pos_for_this_action = [] # POs involved in the specific action this week
        action_po_quantity_sum = 0
        action_description = None
        flags = []

        # --- Calculate demand for the next 6 weeks from the current snapshot's perspective ---
        demand_next_6_weeks = 0
        current_snapshot_demand_view = demand_rows[demand_rows['Snapshot'] == snapshot]
        if not current_snapshot_demand_view.empty:
            for k_future_week_idx in range(1, 7): # Weeks t+1 to t+6
                future_week_col_name = f"WW{week_num + k_future_week_idx}"
                if future_week_col_name in current_snapshot_demand_view.columns:
                    try:
                        future_demand_val = current_snapshot_demand_view[future_week_col_name].iloc[0]
                        # Ensure empty strings or non-convertible values become 0
                        if pd.notna(future_demand_val) and str(future_demand_val).strip() != '':
                            demand_next_6_weeks += int(float(str(future_demand_val))) # float for robustness then int
                        else:
                            demand_next_6_weeks += 0
                    except (ValueError, TypeError):
                        demand_next_6_weeks += 0 # If conversion fails, add 0
        
        # --- Calculate the metric as per user's formula ---
        # Metric: (demand for next 6 weeks - current week's end inventory) / 6
        metric_value = (end_inventory_calc) / (demand_next_6_weeks / 6.0) if demand_next_6_weeks != 0 else (end_inventory_calc)

        # --- POs scheduled for this week (candidates for push-out) ---
        current_week_pos_df_candidates = po_df[
            (po_df['GR WW'] == week_num) & 
            (~po_df['Purchasing Document'].isin(suggested_action_po_docs_globally))
        ]
        current_week_po_docs_from_po_df = [str(doc) for doc in current_week_pos_df_candidates['Purchasing Document']]
        current_week_po_qty_from_po_df = current_week_pos_df_candidates['GR Quantity'].sum()

        # --- Push-out Identification Logic ---
        # Condition: metric_value > 20
        # User's note: This condition implies future demand significantly exceeds current inventory.
        # Pushing out POs in such a scenario is counter-intuitive for typical inventory management if it means reducing incoming supply.
        # Implementing as per user's direct instruction.
        if metric_value > 20:
            if not current_week_pos_df_candidates.empty:
                flags.append(f"Metric ({metric_value:.2f}) > 20. Identifying potential push-out.")
                
                po_to_push_out_row = current_week_pos_df_candidates.iloc[0] # Suggest pushing the first candidate
                doc_to_push = str(po_to_push_out_row['Purchasing Document'])
                qty_to_push = po_to_push_out_row['GR Quantity']

                action = "Suggest Push Out"
                suggested_pos_for_this_action.append(doc_to_push)
                suggested_action_po_docs_globally.add(doc_to_push) # Mark as actioned
                action_po_quantity_sum = qty_to_push 
                
                pushed_to_week_suggestion = week_num + 1 
                action_description = (f"Metric ({metric_value:.2f}) > 20. "
                                      f"Suggest pushing PO {doc_to_push} (Qty {qty_to_push}) "
                                      f"from WW{week_num} to WW{pushed_to_week_suggestion}.")
                flags.append(f"Push-out suggested for PO {doc_to_push}.")
            else:
                flags.append(f"Metric ({metric_value:.2f}) > 20, but no current week POs available to suggest for push-out.")
        
        # --- Pull-in Identification Logic ---
        # Condition: metric_value < lead_time (integer variable from function args)
        elif metric_value < lead_time:
            flags.append(f"Metric ({metric_value:.2f}) < Lead Time ({lead_time}). Identifying potential pull-in.")
            action = "Suggest Pull In" # Tentatively set action

            target_pull_in_quantity = 0
            if end_inventory_calc < 0:
                target_pull_in_quantity = abs(end_inventory_calc)
                flags.append(f"Current inventory ({end_inventory_calc}) is negative. Targeting to cover this deficit ({target_pull_in_quantity}).")
            else:
                # Proactive pull-in: inventory is non-negative, but metric suggests a need.
                # Target pulling in approximately one average week of future demand as a proactive measure.
                if demand_next_6_weeks > 0:
                    avg_one_week_future_demand = demand_next_6_weeks / 6.0
                    target_pull_in_quantity = avg_one_week_future_demand
                    flags.append(f"Proactive pull-in. Targeting ~1 avg future week's demand ({target_pull_in_quantity:.0f}).")
                else: 
                    # No future demand, inventory non-negative, but metric triggered pull-in.
                    # This case (e.g., positive inventory, metric = -inv/6 < lead_time) is complex.
                    # Pulling more when no future demand and positive inventory might be undesirable.
                    # For now, set a nominal small target to see if any PO can be pulled.
                    target_pull_in_quantity = 1 
                    flags.append("Proactive pull-in triggered with no future demand. Nominal target (1 unit) to identify next PO.")
            
            if target_pull_in_quantity > 0:
                future_po_candidates_for_pull = po_df[
                    (po_df['GR WW'] > week_num) & 
                    (~po_df['Purchasing Document'].isin(suggested_action_po_docs_globally))
                ].sort_values(by=['GR WW', 'Order WW'])

                qty_identified_for_pull_in_total = 0
                pulled_in_details_desc_list = []
                # Make a copy of target for loop modification
                remaining_qty_to_target_for_pull = target_pull_in_quantity 

                for _, po_row in future_po_candidates_for_pull.iterrows():
                    # If we've met or exceeded the target, and have identified at least one PO
                    if remaining_qty_to_target_for_pull <= 0 and qty_identified_for_pull_in_total > 0:
                        break 

                    po_doc = str(po_row['Purchasing Document'])
                    po_qty = po_row['GR Quantity']
                    original_gr_ww = po_row['GR WW']

                    qty_identified_for_pull_in_total += po_qty
                    suggested_pos_for_this_action.append(po_doc)
                    suggested_action_po_docs_globally.add(po_doc) # Mark as actioned
                    pulled_in_details_desc_list.append(f"{po_doc} (Qty {po_qty} from WW{original_gr_ww})")
                    
                    remaining_qty_to_target_for_pull -= po_qty
                
                if qty_identified_for_pull_in_total > 0:
                    action_po_quantity_sum = qty_identified_for_pull_in_total
                    action_description = (f"Metric ({metric_value:.2f}) < Lead Time ({lead_time}). "
                                          f"Suggest pulling to WW{week_num}: {'; '.join(pulled_in_details_desc_list)}.")
                    if remaining_qty_to_target_for_pull > 0: # Means we couldn't cover the full target
                         flags.append(f"Suggested pull-in may be insufficient for the calculated target need of {target_pull_in_quantity:.0f}.")
                    else:
                         flags.append(f"Suggested pull-in likely covers/exceeds calculated target need of {target_pull_in_quantity:.0f}.")
                else:
                    flags.append(f"Metric condition for pull-in met (target {target_pull_in_quantity:.0f}), but no available POs identified.")
                    action = "No Action" # Revert action if no POs found
            else: # target_pull_in_quantity was 0 or less.
                flags.append(f"Metric condition for pull-in met, but calculated target quantity for pull-in is zero or less. No specific POs targeted.")
                action = "No Action" # Revert action
        else: # Neither Push Out nor Pull In conditions met based on the metric
            action = "No Action" 
            flags.append(f"Metric ({metric_value:.2f}) did not trigger push-out (>20) or pull-in (<{lead_time}). Conditions OK.")

        # Ensure 'OK' is the primary flag if no specific action/warning arose.
        if action == "No Action" and not any("risk" in f.lower() or "negative" in f.lower() or "insufficient" in f.lower() or "push-out suggested" in f.lower() or "pull-in suggested" in f.lower() for f in flags):
            # If it's "No Action" and no other critical flags, simplify to "OK" or the metric status.
            if not (metric_value > 20 or metric_value < lead_time): # If metric is in the "OK" band
                 flags = [f"Metric ({metric_value:.2f}) is within acceptable range. OK."]
            # else, the existing flags about metric conditions not being met are fine.


        final_flags = [f for f in flags if f] # Remove empty strings
        if not final_flags: # Should not happen if logic above is complete
             final_flags.append("OK")


        results.append({
            'Snapshot Week': snapshot,
            'Start Inventory (Waterfall Ref)': start_inventory_waterfall_ref,
            'Start Inventory (Calc)': opening_inventory_calc,
            'Demand (Waterfall)': demand_waterfall,
            'Supply (Waterfall)': supply_waterfall,
            'Demand Next 6 Weeks (View)': demand_next_6_weeks, # Added for transparency
            'WOS': round(metric_value, 2), # Added for transparency
            'PO Docs Scheduled This Week (po_df)': ", ".join(current_week_po_docs_from_po_df) if current_week_po_docs_from_po_df else None,
            'PO Qty Scheduled This Week (po_df)': current_week_po_qty_from_po_df,
            'End Inventory (Calc)': end_inventory_calc,
            'Suggested POs for Action': ", ".join(suggested_pos_for_this_action) if suggested_pos_for_this_action else None,
            'Quantity of POs in Suggested Action': action_po_quantity_sum,
            'Suggested Action Detail': action_description,
            'Flags': ", ".join(final_flags),
            'Identified Action': action
        })

        # Update calculated inventory for the next iteration
        current_inventory_calc = end_inventory_calc
        
    return pd.DataFrame(results)

def analyze_discrepancy_scen_7(row):
    planned_supply = row['Supply (Waterfall)']
    incoming_po_gr = row['GR Quantity']
    po_docs = ", ".join(str(int(x)) for x in row['Purchasing Document']) if row['Purchasing Document'] else "None"

    discrepancy_flag = False
    discrepancy_detail = "No discrepancy."

    abs_diff = abs(planned_supply - incoming_po_gr)

    if planned_supply == 0 and incoming_po_gr != 0:
        discrepancy_flag = True
        discrepancy_detail = (f"Discrepancy: Planned Supply is 0, but Goods Receipt (GR) from PO(s) is {incoming_po_gr} "
                              f"(difference of {abs_diff}). Affected PO(s): {po_docs}.")
    elif planned_supply != 0 and incoming_po_gr == 0:
        discrepancy_flag = True
        discrepancy_detail = (f"Discrepancy: Planned Supply is {planned_supply}, but Goods Receipt (GR) from PO(s) is 0 "
                              f"(difference of {abs_diff}). Expected PO(s): {po_docs}.")
    elif planned_supply != 0 and incoming_po_gr != 0 and planned_supply != incoming_po_gr:
        discrepancy_flag = True
        difference = incoming_po_gr - planned_supply
        if difference > 0:
            discrepancy_detail = (f"Discrepancy: Planned Supply ({planned_supply}) is less than Goods Receipt (GR) from PO(s) "
                                  f"({incoming_po_gr}) by {abs_diff}. Affected PO(s): {po_docs}.")
        else:
            discrepancy_detail = (f"Discrepancy: Planned Supply ({planned_supply}) is greater than Goods Receipt (GR) from PO(s) "
                                  f"({incoming_po_gr}) by {abs_diff}. Affected PO(s): {po_docs}.")
    elif planned_supply == 0 and incoming_po_gr == 0:
        pass

    if not discrepancy_flag and po_docs != "None":
        discrepancy_detail = f"Planned Supply matches Goods Receipt from PO(s). PO(s) for this week: {po_docs}."
    elif not discrepancy_flag and po_docs == "None":
        discrepancy_detail = "No discrepancy. No POs scheduled for GR this week."

    return pd.Series([discrepancy_flag, discrepancy_detail, abs_diff])